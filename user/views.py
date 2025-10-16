from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import User
from .serializers import UserSerializer
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter, OrderingFilter
from cms.utils.pagination import CustomPageNumberPagination
from cms.utils.filter import UserFilter
from django_filters import rest_framework as filters
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class UserViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    permission_classes = [AllowAny]  # Temporarily allow for testing
    pagination_class = CustomPageNumberPagination
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = UserFilter
    search_fields = ['username', 'first_name', 'last_name', 'email']  # Fields to search in
    ordering_fields = ['username', 'first_name', 'last_name', 'email', 'role', 'is_active', 'date_joined']
    ordering = ['username']

    def get_queryset(self):
        if self.action == 'retrieve':
            return User.objects.filter(id=self.kwargs['pk'])

        elif self.action == 'list':
            return User.objects.all()  # Return all users for list to enable proper filtering
        return User.objects.all()  # For other actions like update, delete, etc.

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculate counts from filtered queryset (respects search and filters)
        active_count = queryset.filter(is_active=True).count()
        inactive_count = queryset.filter(is_active=False).count()
        managers_count = queryset.filter(role=User.MANAGER).count()
        masters_count = queryset.filter(role=User.MASTER).count()
        
        # Paginate
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['total_active_count'] = active_count
            response.data['total_inactive_count'] = inactive_count
            response.data['total_managers_count'] = managers_count
            response.data['total_masters_count'] = masters_count
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'count': queryset.count(),
            'total_active_count': active_count,
            'total_inactive_count': inactive_count,
            'total_managers_count': managers_count,
            'total_masters_count': masters_count
        })

    def perform_create(self, serializer):
        user = serializer.save()
        password = self.request.data.get('password')
        if password:
            user.set_password(password)
            user.save()

    def perform_update(self, serializer):
        user = serializer.save()
        password = self.request.data.get('password')
        if password:
            user.set_password(password)
            user.save()

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def details(self, request):
        """
        Custom action to get the details of the currently authenticated user.
        Accessible at /api/users/details/ with the token in the Authorization header.
        """
        user = request.user
        serializer = UserSerializer(user)
        return Response(serializer.data)


class UserExportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        # Apply filters using the same filter class as UserViewSet
        filter_instance = UserFilter(request.GET, queryset=User.objects.all())
        queryset = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        # Apply ordering same as UserViewSet
        queryset = queryset.order_by('username')

        # Limit results for performance
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)
        users = queryset[:export_limit]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Users"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Username', 'First Name', 'Last Name', 'Email', 'Role', 'Is Active', 'Is Staff', 'Date Joined', 'Last Login']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row, user in enumerate(users, 2):
            data = [
                user.username,
                user.first_name,
                user.last_name,
                user.email,
                user.get_role_display(),
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''
            ]

            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        workbook.save(response)
        return response

