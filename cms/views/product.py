from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from cms.models.product import Product, ProductOption, ProductVariant, ProductVariantImage, Collection, ProductLinkVariant, ProductPriceHistory, ProductVariantCustomField, ProductSizeChartValue, ComboProduct, ComboProductItem
from cms.models.product_image import ProductImage
from cms.models.facility import Facility, FacilityInventory, Cluster
from cms.models.category import Brand, Category
from cms.models.setting import CustomField, SizeChart, SizeMeasurement, AttributeValue

import csv
import pandas as pd
from cms.serializers.product import (
    ProductListSerializer,
    ProductDetailSerializer,
    CollectionSerializer,
    CollectionListSerializer,
    ProductVariantListSerializer,
    ProductStatusUpdateSerializer,
    SingleProductSerializer,
    BulkProductSerializer,
    ProductExportSerializer,
    ProductWithClusterPricingSerializer,
    ProductWithFacilityPricingSerializer,
    ClusterPriceUpdateSerializer,
    ProductPriceHistorySerializer,
    SmartBrandProductSerializer,
    ComboProductListSerializer,
    ComboProductCreateSerializer
)
from cms.utils.filter import (
    ProductVariantFilter,
    ProductFilter,
    CollectionFilter,
    ComboProductFilter
)
from cms.utils.pagination import CustomPageNumberPagination
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
from django.db.models import Q, Max
from django.http import HttpResponse
from django_filters import rest_framework as filters
import time


class CollectionViewSet(viewsets.ModelViewSet):
    queryset = Collection.objects.all()
    # serializer_class = CollectionSerializer
    permission_classes  = [IsAuthenticated, DjangoModelPermissions]
    pagination_class    = CustomPageNumberPagination
    filter_backends     = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class     = CollectionFilter
    search_fields       = ['name', 'description']
    ordering_fields     = ['name', 'is_active', 'creation_date', 'updation_date']
    ordering            = ['name']

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return CollectionListSerializer
        else:
            return CollectionSerializer


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends  = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class  = ProductFilter
    search_fields    = ['name', 'sku', 'description', 'tags']
    ordering_fields  = ['name', 'category__name', 'brand__name', 'is_active', 'creation_date', 'updation_date']
    ordering         = ['creation_date']

    # def get_queryset(self):
    #     current_user = self.request.user

    #     # Filter products to only show those with non-rejected variants
    #     # Exclude products that have no variants or only rejected variants
    #     queryset = Product.objects.filter(
    #         variants__is_rejected=False
    #     ).distinct()

    #     if current_user.role == 'manager':
    #         managed_facilities = Facility.objects.filter(managers=current_user)
    #         product_variant_ids = FacilityInventory.objects.filter(
    #             facility__in=managed_facilities
    #         ).values_list('product_variant', flat=True)
    #         return queryset.filter(variants__in=product_variant_ids).distinct()

    #     return queryset
    
    def get_queryset(self):
        current_user = self.request.user
        show_rejected = self.request.query_params.get('rejected') == 'true'

        # If rejected=true, filter products with rejected variants
        if show_rejected:
            queryset = Product.objects.filter(
                variants__is_rejected=True
            ).distinct()
        else:
            queryset = Product.objects.filter(
                variants__is_rejected=False
            ).distinct()

        # if current_user.role == 'manager':
        #     managed_facilities = Facility.objects.filter(managers=current_user)
        #     product_variant_ids = FacilityInventory.objects.filter(
        #         facility__in=managed_facilities
        #     ).values_list('product_variant', flat=True)
        #     queryset = queryset.filter(variants__in=product_variant_ids).distinct()

        return queryset


    def get_serializer_class(self):
        if self.action == 'list':
            # Use a lighter serializer for list view to avoid performance issues
            from cms.serializers.product import ProductViewSerializer
            return ProductViewSerializer
        elif self.action == 'retrieve':
            return ProductListSerializer
        return ProductDetailSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        
        product_options = validated_data.pop('options', [])
        product_variants = validated_data.pop('variants', [])
        associated_facility_ids = validated_data.pop('facilities', [])
        collection_ids = validated_data.pop('collections', [])
        linked_variant_ids = validated_data.pop('linked_variants', [])

        # Creating the product object with user tracking
        new_product = Product.objects.create(
            created_by=request.user,
            updated_by=request.user,
            **validated_data
        )

        # Creating product options if any
        for option_data in product_options:
            ProductOption.objects.create(product=new_product, **option_data)

        # Handle pack variants with link-based assignment and custom fields
        created_variants = self._create_variants_with_pack_support(new_product, product_variants)
        
        new_product.variants.set(created_variants)

        # Handling facility inventory creation for associated facilities
        for facility_id in associated_facility_ids:
            try:
                facility = Facility.objects.get(pk=facility_id)
            except Facility.DoesNotExist:
                continue
            missing_inventory = [
                FacilityInventory(
                    facility=facility,
                    product_variant=product_variant,
                    stock=0,
                    base_price=product_variant.base_price or 0.0,
                    mrp=product_variant.mrp or 0.0,
                    selling_price=product_variant.selling_price if hasattr(product_variant, 'selling_price') else product_variant.base_price or 0.0
                )
                for product_variant in created_variants
                if not FacilityInventory.objects.filter(facility=facility, product_variant=product_variant).exists()
            ]
            FacilityInventory.objects.bulk_create(missing_inventory)

        # Managing collections associated with the new product
        for collection_id in collection_ids:
            try:
                collection = Collection.objects.get(pk=collection_id)
                new_product.collections.add(collection)
            except Collection.DoesNotExist:
                continue

        # Linking products
        for linked_variant_id in linked_variant_ids:
            try:
                linked_variant = ProductVariant.objects.get(pk=linked_variant_id)
                ProductLinkVariant.objects.create(product=new_product, linked_variant=linked_variant)
            except ProductVariant.DoesNotExist:
                continue

        # Returning the created product details
        product_details = ProductDetailSerializer(new_product, context={'request': request})
        return Response(product_details.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        product = self.get_object()
        serializer = self.get_serializer(product, data=request.data, partial=kwargs.pop('partial', False))
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        # Update direct fields
        direct_fields = [
            'category', 'brand', 'name', 'sku', 'description', 'tags', 'is_active'
        ]
        for field in direct_fields:
            if field in validated_data:
                setattr(product, field, validated_data[field])
        
        product.updated_by = request.user
        product.save()

        # Handling nested fields
        product_options = validated_data.pop('options', [])
        product_variants = validated_data.pop('variants', [])
        associated_facility_ids = validated_data.pop('facilities', [])
        collection_ids = validated_data.pop('collections', [])
        linked_variant_ids = validated_data.pop('linked_variants', [])

        # Update product options
        existing_option_ids = {option.id: option for option in product.options.all()}

        for option_data in product_options:
            option_id = option_data.get('id', None)
            if option_id:
                existing_option = existing_option_ids.get(option_id)
                if existing_option:
                    for field, value in option_data.items():
                        setattr(existing_option, field, value)
                    existing_option.save()
            else:
                ProductOption.objects.create(product=product, **option_data)

        # Remove options that are not in the request
        options_to_remove = set(existing_option_ids.keys()) - {option_data.get('id') for option_data in product_options}
        ProductOption.objects.filter(id__in=options_to_remove, product=product).delete()

        # Handle variants with pack support and custom fields in update
        self._update_variants_with_pack_support(product, product_variants)

        # Handling linked products
        existing_linked_variants = {link.linked_variant.id: link for link in product.linked_variants.all()}
        for linked_variant_id in linked_variant_ids:
            try:
                linked_variant = ProductVariant.objects.get(pk=linked_variant_id)
                if linked_variant_id in existing_linked_variants:
                    existing_linked_variants[linked_variant_id].save()
                else:
                    ProductLinkVariant.objects.create(product=product, linked_variant=linked_variant)
            except ProductVariant.DoesNotExist:
                continue

        # Handling facility inventory updates
        for facility_id in associated_facility_ids:
            try:
                facility = Facility.objects.get(pk=facility_id)
            except Facility.DoesNotExist:
                continue
            missing_inventory = [
                FacilityInventory(
                    facility=facility,
                    product_variant=product_variant,
                    stock=0,
                    base_price=product_variant.base_price or 0.0,
                    mrp=product_variant.mrp or 0.0,
                    selling_price=product_variant.selling_price if hasattr(product_variant, 'selling_price') else product_variant.base_price or 0.0
                )
                for product_variant in product.variants.all()
                if not FacilityInventory.objects.filter(facility=facility, product_variant=product_variant).exists()
            ]
            FacilityInventory.objects.bulk_create(missing_inventory)

        # Managing product collections
        product.collections.clear()
        for collection_id in collection_ids:
            try:
                collection = Collection.objects.get(pk=collection_id)
                product.collections.add(collection)
            except Collection.DoesNotExist:
                continue

        # Returning the updated product details
        product_details = ProductDetailSerializer(product, context={'request': request})
        return Response(product_details.data, status=status.HTTP_200_OK)

    def _handle_variant_images(self, variant, variant_images):
        """Helper method to handle variant images during create/update"""
        existing_images = {img.id: img for img in variant.images.all()}
        new_image_ids = []
        existing_image_ids_kept = []

        for image_data in variant_images:
            image_id = image_data.get('id', None)
            if image_id and image_id in existing_images:
                existing_image = existing_images[image_id]
                for field, value in image_data.items():
                    setattr(existing_image, field, value)
                existing_image.save()
                existing_image_ids_kept.append(existing_image.id)
            else:
                new_image = ProductVariantImage.objects.create(product_variant=variant, **image_data)
                new_image_ids.append(new_image.id)

        image_ids_to_keep = existing_image_ids_kept + new_image_ids
        variant.images.exclude(id__in=image_ids_to_keep).delete()

    def _handle_variant_custom_fields(self, variant, custom_fields_data):
        """Handle custom field values for a variant"""
        if not custom_fields_data:
            return
            
        # Get existing custom field values for this variant
        existing_custom_values = {
            cv.custom_field.id: cv 
            for cv in variant.custom_field_values.all()
        }
        
        processed_field_ids = set()
        
        for field_data in custom_fields_data:
            field_id = field_data.get('field_id')
            field_value = field_data.get('value', '')
            
            if not field_id:
                continue
                
            try:
                custom_field = CustomField.objects.get(id=field_id)
                processed_field_ids.add(field_id)
                
                if field_id in existing_custom_values:
                    # Update existing custom field value
                    existing_custom_values[field_id].value = field_value
                    existing_custom_values[field_id].save()
                else:
                    # Create new custom field value
                    ProductVariantCustomField.objects.create(
                        product_variant=variant,
                        custom_field=custom_field,
                        value=field_value
                    )
            except CustomField.DoesNotExist:
                print(f"Custom field with ID {field_id} not found")
                continue
        
        # Remove custom field values that are not in the request
        fields_to_remove = set(existing_custom_values.keys()) - processed_field_ids
        if fields_to_remove:
            ProductVariantCustomField.objects.filter(
                product_variant=variant,
                custom_field_id__in=fields_to_remove
            ).delete()

    def _create_variants_with_pack_support(self, product, product_variants):
        """Create variants with pack support using link field and custom fields"""
        created_variants = []
        
        # Group variants by link value
        variants_by_link = {}
        for variant_data in product_variants:
            link = variant_data.get('link')
            if link not in variants_by_link:
                variants_by_link[link] = []
            variants_by_link[link].append(variant_data)
                
        # Store created single variants by link for reference
        single_variants_by_link = {}
        
        # First pass: Create single unit variants
        for link, variants_in_group in variants_by_link.items():
            for variant_data in variants_in_group:
                variant_images = variant_data.pop('images', [])
                custom_fields = variant_data.pop('custom_fields', [])
                size_chart_values = variant_data.pop('size_chart_values', [])
                is_pack = str(variant_data.get('is_pack', 'false')).lower() == 'true'
                pack_qty = variant_data.get('pack_qty', 1) or 1

                if not is_pack or pack_qty == 1:
                    clean_variant_data = self._clean_variant_data_for_single(variant_data)

                    product_variant = ProductVariant.objects.create(product=product, **clean_variant_data)
                    created_variants.append(product_variant)

                    # Store for pack variants to reference
                    single_variants_by_link[link] = product_variant

                    # Create variant images
                    for image_data in variant_images:
                        ProductVariantImage.objects.create(product_variant=product_variant, **image_data)

                    # Handle custom fields
                    self._handle_variant_custom_fields(product_variant, custom_fields)

                    # Handle size chart values
                    if size_chart_values:
                        handle_product_size_chart(product_variant, {'size_chart_values': size_chart_values})
                
        # Second pass: Create pack variants with base_variant references
        for link, variants_in_group in variants_by_link.items():
            for variant_data in variants_in_group:
                variant_images = variant_data.pop('images', [])
                custom_fields = variant_data.pop('custom_fields', [])
                size_chart_values = variant_data.pop('size_chart_values', [])
                is_pack = str(variant_data.get('is_pack', 'false')).lower() == 'true'
                pack_qty = variant_data.get('pack_qty', 1) or 1

                if is_pack and pack_qty > 1:
                    # Find the base variant with same link
                    base_variant = single_variants_by_link.get(link)

                    if base_variant:
                        clean_variant_data = self._clean_variant_data_for_pack(variant_data, base_variant)

                        product_variant = ProductVariant.objects.create(product=product, **clean_variant_data)
                        created_variants.append(product_variant)

                        # Create variant images
                        for image_data in variant_images:
                            ProductVariantImage.objects.create(product_variant=product_variant, **image_data)

                        # Handle custom fields
                        self._handle_variant_custom_fields(product_variant, custom_fields)

                        # Handle size chart values
                        if size_chart_values:
                            handle_product_size_chart(product_variant, {'size_chart_values': size_chart_values})
                    else:
                        print(f"  ❌ WARNING: No base variant found for link {link}")
        
        return created_variants

    def _update_variants_with_pack_support(self, product, product_variants):
        """Update variants with pack support and custom fields"""
        existing_variant_ids = {variant.id: variant for variant in product.variants.all()}
        new_variant_ids = []
        request_variant_ids = [variant_data.get('id') for variant_data in product_variants if variant_data.get('id')]
        
        # Separate existing and new variants
        existing_variants = []
        new_variants = []
        
        for variant_data in product_variants:
            if variant_data.get('id'):
                existing_variants.append(variant_data)
            else:
                new_variants.append(variant_data)
        
        # Process existing variants
        for variant_data in existing_variants:
            variant_images = variant_data.pop('images', [])
            custom_fields = variant_data.pop('custom_fields', [])
            size_chart_values = variant_data.pop('size_chart_values', [])
            variant_id = variant_data.get('id')

            if variant_id in existing_variant_ids:
                existing_variant = existing_variant_ids[variant_id]

                clean_variant_data = self._clean_variant_data_for_existing(variant_data)

                for field, value in clean_variant_data.items():
                    setattr(existing_variant, field, value)
                existing_variant.save()

                self._handle_variant_images(existing_variant, variant_images)
                self._handle_variant_custom_fields(existing_variant, custom_fields)

                # Handle size chart values for existing variants
                if size_chart_values:
                    handle_product_size_chart(existing_variant, {'size_chart_values': size_chart_values})

                new_variant_ids.append(variant_id)
        
        # Process new variants using the same logic as create
        if new_variants:
            variants_by_link = {}
            for variant_data in new_variants:
                link = variant_data.get('link')
                if link not in variants_by_link:
                    variants_by_link[link] = []
                variants_by_link[link].append(variant_data)
            
            single_variants_by_link = {}
            
            # First: Create new single variants
            for link, variants_in_group in variants_by_link.items():
                for variant_data in variants_in_group:
                    variant_images = variant_data.pop('images', [])
                    custom_fields = variant_data.pop('custom_fields', [])
                    size_chart_values = variant_data.pop('size_chart_values', [])
                    is_pack = str(variant_data.get('is_pack', 'false')).lower() == 'true'
                    pack_qty = variant_data.get('pack_qty', 1) or 1

                    if not is_pack or pack_qty == 1:
                        clean_variant_data = self._clean_variant_data_for_single(variant_data)
                        new_variant = ProductVariant.objects.create(product=product, **clean_variant_data)

                        self._handle_variant_images(new_variant, variant_images)
                        self._handle_variant_custom_fields(new_variant, custom_fields)

                        # Handle size chart values for new single variants
                        if size_chart_values:
                            handle_product_size_chart(new_variant, {'size_chart_values': size_chart_values})

                        single_variants_by_link[link] = new_variant
                        new_variant_ids.append(new_variant.id)
            
            # Second: Create new pack variants
            for link, variants_in_group in variants_by_link.items():
                for variant_data in variants_in_group:
                    variant_images = variant_data.pop('images', [])
                    custom_fields = variant_data.pop('custom_fields', [])
                    size_chart_values = variant_data.pop('size_chart_values', [])
                    is_pack = str(variant_data.get('is_pack', 'false')).lower() == 'true'
                    pack_qty = variant_data.get('pack_qty', 1) or 1

                    if is_pack and pack_qty > 1:
                        base_variant = single_variants_by_link.get(link)

                        if not base_variant:
                            for existing_data in existing_variants:
                                if existing_data.get('link') == link:
                                    existing_id = existing_data.get('id')
                                    base_variant = existing_variant_ids.get(existing_id)
                                    break

                        if base_variant:
                            clean_variant_data = self._clean_variant_data_for_pack(variant_data, base_variant)
                            new_variant = ProductVariant.objects.create(product=product, **clean_variant_data)

                            self._handle_variant_images(new_variant, variant_images)
                            self._handle_variant_custom_fields(new_variant, custom_fields)

                            # Handle size chart values for new pack variants
                            if size_chart_values:
                                handle_product_size_chart(new_variant, {'size_chart_values': size_chart_values})

                            new_variant_ids.append(new_variant.id)
                        else:
                            print(f"  ❌ WARNING: No base variant found for link {link}")
        
        # Remove variants that are not in the request
        variants_to_remove = set(existing_variant_ids.keys()) - set(request_variant_ids)
        if variants_to_remove:
            product.variants.filter(id__in=variants_to_remove).delete()
        
        # Set the updated variants
        all_variant_ids = set(new_variant_ids + request_variant_ids)
        product.variants.set(ProductVariant.objects.filter(id__in=all_variant_ids))

    def _clean_variant_data_for_single(self, variant_data):
        """Clean and prepare variant data for single unit variant"""
        clean_data = variant_data.copy()
        
        # Remove fields that shouldn't be passed to create/update
        clean_data.pop('id', None)
        clean_data.pop('link', None)
        clean_data.pop('images', None)
        clean_data.pop('custom_fields', None)  # Remove custom fields from variant data
        
        # Ensure single unit properties
        clean_data['is_pack'] = False
        clean_data['pack_qty'] = 1
        clean_data['pack_variant'] = None
        
        return clean_data

    def _clean_variant_data_for_pack(self, variant_data, base_variant):
        """Clean and prepare variant data for pack variant"""
        clean_data = variant_data.copy()
        
        # Remove fields that shouldn't be passed to create/update
        clean_data.pop('id', None)
        clean_data.pop('link', None)
        clean_data.pop('images', None)
        clean_data.pop('custom_fields', None)  # Remove custom fields from variant data
        
        # Set pack properties
        pack_qty = variant_data.get('pack_qty', 1) or 1
        clean_data['is_pack'] = True
        clean_data['pack_qty'] = pack_qty
        clean_data['pack_variant'] = base_variant
        
        return clean_data

    def _clean_variant_data_for_existing(self, variant_data):
        """Clean variant data for existing variants - preserves pack_variant field"""
        clean_data = variant_data.copy()
        
        # Remove fields that shouldn't be passed to create/update
        clean_data.pop('id', None)
        clean_data.pop('link', None)
        clean_data.pop('images', None)
        clean_data.pop('custom_fields', None)  # Remove custom fields from variant data
        
        return clean_data
    
    
class ProductStatusUpdateView(APIView):
    queryset = Product.objects.all()
    """
    A view to update the status (boolean field) of a product.
    """
    permission_classes = [IsAuthenticated, DjangoModelPermissions]

    def patch(self, request, *args, **kwargs):
        # Get the product by ID
        product_id = kwargs.get('product_id')
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

        # Deserialize the request data to get the status
        serializer = ProductStatusUpdateSerializer(data=request.data)
        if serializer.is_valid():
            # Update the product's status
            product.is_active = serializer.validated_data['is_active']
            product.save()

            return Response({"status": "Product status updated successfully."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductVariantViewSet(viewsets.ModelViewSet):
    queryset = ProductVariant.objects.all()
    permission_classes  = [IsAuthenticated, DjangoModelPermissions]
    pagination_class    = CustomPageNumberPagination
    filter_backends     = (filters.DjangoFilterBackend, SearchFilter)
    filterset_class     = ProductVariantFilter
    search_fields       = ['name', 'sku', 'product__name']

    def get_serializer_class(self):
        return ProductVariantListSerializer

class ProductListViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()

    def get_serializer_class(self):
        return ProductListSerializer

# VIEW – skip name-dupes, create only genuinely new ones with smart brand handling

class BulkCreateProductsView(APIView):
    """
    POST /api/products/bulk-create/
    Body: JSON array of product objects for CREATION ONLY (no SKUs expected)
    Brand field accepts either ID (numeric) or name (text).
    SKUs will be auto-generated for all variants.

    Custom fields can be included in variant data under 'custom_fields' key in two formats:
    1. Array format: [{"field_id": 13, "value": "Yes"}]
    2. Dict format: {"field_name": "value"} (legacy)
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        from django.db import transaction
        import time

        start_time = time.time()
        print(f"🚀 BULK CREATE MODE: Creating {len(request.data)} new products...")

        try:
            with transaction.atomic():
                # Step 1: Validate that no SKUs are provided (creation only)
                validation_start = time.time()
                for idx, item in enumerate(request.data):
                    if isinstance(item, dict) and 'variants' in item:
                        for variant_data in item['variants']:
                            if isinstance(variant_data, dict) and variant_data.get('sku'):
                                return Response({
                                    'error': f'SKUs are not allowed in creation endpoint. Found SKU "{variant_data.get("sku")}" in product {idx + 1}.',
                                    'message': 'Use the bulk-update endpoint (PUT) for updating existing products with SKUs.'
                                }, status=400)

                print(f"⚡ SKU validation: {round((time.time() - validation_start) * 1000, 1)}ms - No SKUs found (correct for creation)")

                # Step 3: Collect EAN numbers for validation
                ean_start = time.time()
                all_ean_numbers = []
                for item in request.data:
                    if isinstance(item, dict) and 'variants' in item:
                        for variant_data in item['variants']:
                            if isinstance(variant_data, dict) and variant_data.get('ean_number'):
                                all_ean_numbers.append(variant_data['ean_number'])

                # Fast EAN validation (optional - can be disabled for maximum speed)
                ean_validation_results = {}
                ean_rejected_products = []

                if all_ean_numbers and len(all_ean_numbers) <= 200:  # Only validate if reasonable amount
                    try:
                        import os, requests, json
                        gs1_url = os.environ.get("GS1_API_URL")
                        gs1_token = os.environ.get("GS1_API_TOKEN", "")

                        if gs1_url and gs1_token:
                            session = requests.Session()
                            session.headers.update({'Authorization': f'Bearer {gs1_token}'})

                            # Process in single batch for speed
                            gtin_param = json.dumps(all_ean_numbers)
                            response = session.get(
                                gs1_url,
                                params={'gtin': gtin_param, 'status': 'published'},
                                timeout=3  # Very short timeout
                            )

                            if response.status_code == 200:
                                data = response.json()
                                if data.get('status') and data.get('items'):
                                    for item in data['items']:
                                        ean = item.get('gtin')
                                        if ean:
                                            ean_validation_results[ean] = {
                                                'hsn_code': item.get('hs_code'),
                                                'tax': item.get('tax_rate', 0) or item.get('igst', 0),
                                                'cgst': item.get('cgst', 0),
                                                'sgst': item.get('sgst', 0),
                                                'igst': item.get('igst', 0),
                                                'cess': item.get('cess', 0),
                                                'is_valid': True
                                            }

                            session.close()
                    except Exception:
                        pass  # Continue without EAN validation

                # Mark unvalidated EANs
                for ean in all_ean_numbers:
                    if ean not in ean_validation_results:
                        ean_validation_results[ean] = {'is_valid': False}

                print(f"⚡ EAN validation: {round((time.time() - ean_start) * 1000, 1)}ms - {len(ean_validation_results)} EANs checked")

                # Step 4: Create products processing
                process_start = time.time()
                created_products = []
                created_count = 0
                failed_count = 0
                failed_products = []
                ean_rejected_products = []

                for idx, item in enumerate(request.data):
                    try:
                        # Detailed validation with specific error messages
                        validation_errors = []

                        if not isinstance(item, dict):
                            validation_errors.append("Item must be a dictionary/object")
                        else:
                            if not item.get('name'):
                                validation_errors.append("Product name is required")
                            if not item.get('category'):
                                validation_errors.append("Product category is required")
                            else:
                                # Validate category exists
                                try:
                                    Category.objects.get(id=item.get('category'))
                                except Category.DoesNotExist:
                                    validation_errors.append(f"Category with ID {item.get('category')} does not exist")

                            # Variants are optional - create default if not provided
                            if item.get('variants') is not None:
                                if not isinstance(item.get('variants'), list):
                                    validation_errors.append("Variants must be an array")
                                elif len(item.get('variants', [])) == 0:
                                    validation_errors.append("If variants provided, at least one variant is required")

                            # Validate brand field
                            brand = item.get('brand')
                            if brand is not None and not isinstance(brand, (int, float)):
                                validation_errors.append(f"Brand must be a number (ID), got '{brand}'")

                            # Validate each variant if provided
                            variants = item.get('variants', [])
                            for v_idx, variant in enumerate(variants):
                                if not isinstance(variant, dict):
                                    validation_errors.append(f"Variant {v_idx + 1} must be a dictionary/object")
                                else:
                                    # SKU will be auto-generated, no need to validate for creation
                                    if variant.get('base_price') is not None and not isinstance(variant.get('base_price'), (int, float)):
                                        validation_errors.append(f"Variant {v_idx + 1}: base_price must be a number")
                                    if variant.get('mrp') is not None and not isinstance(variant.get('mrp'), (int, float)):
                                        validation_errors.append(f"Variant {v_idx + 1}: mrp must be a number")
                                    if variant.get('selling_price') is not None and not isinstance(variant.get('selling_price'), (int, float)):
                                        validation_errors.append(f"Variant {v_idx + 1}: selling_price must be a number")

                        if validation_errors:
                            failed_count += 1
                            failed_products.append({
                                'index': idx,
                                'product_name': item.get('name', 'Unknown') if isinstance(item, dict) else 'Unknown',
                                'error': '; '.join(validation_errors),
                                'original_data': item
                            })
                            continue

                        # Handle variants - create default if none provided
                        variants = item.get('variants', [])
                        if not variants:
                            # Create default variant when no variants provided
                            default_variant = {
                                'name': item['name'],  # Use product name as default variant name
                                'sku': None,  # Will be auto-generated
                                'base_price': 0,
                                'mrp': 0,
                                'selling_price': 0
                            }
                            variants = [default_variant]
                            item['variants'] = variants

                        # Create new product
                        try:
                            serializer = SmartBrandProductSerializer(data=item)
                            if serializer.is_valid():
                                result = serializer.save()

                                if isinstance(result, dict):
                                    # Handle serializer result that includes metadata
                                    product = result.get('product')
                                    ean_rejected_variants = result.get('ean_rejected_products', [])

                                    if product:
                                        product.created_by = request.user
                                        product.save()

                                        # Get all variants and separate accepted/rejected
                                        all_variants = product.variants.all()
                                        accepted_variants = [v for v in all_variants if not v.is_rejected]
                                        rejected_variants = [v for v in all_variants if v.is_rejected]

                                        print(f"Debug - Product {product.name}: total_variants={len(all_variants)}, accepted={len(accepted_variants)}, rejected={len(rejected_variants)}")
                                        for v in all_variants:
                                            print(f"  Variant {v.name}: is_rejected={v.is_rejected}")

                                        # Only include in created_products if product has at least one accepted variant
                                        if accepted_variants:
                                            created_products.append({
                                                'product_id': product.id,
                                                'product_name': product.name,
                                                'product_sku': product.sku,
                                                'total_variants': len(all_variants),
                                                'accepted_variants_count': len(accepted_variants),
                                                'rejected_variants_count': len(rejected_variants),
                                                'created_variants': [
                                                    {
                                                        'variant_id': v.id,
                                                        'variant_name': v.name,
                                                        'variant_sku': v.sku,
                                                        'is_rejected': v.is_rejected
                                                    }
                                                    for v in accepted_variants  # Only show accepted variants
                                                ]
                                            })
                                            created_count += 1

                                        # Track EAN rejected variants
                                        if ean_rejected_variants:
                                            for rejected_variant in ean_rejected_variants:
                                                ean_rejected_products.append({
                                                    'product_id': product.id,
                                                    'product_name': product.name,
                                                    'product_sku': product.sku,
                                                    'variant_name': rejected_variant.get('name'),
                                                    'ean_number': rejected_variant.get('ean_number'),
                                                    'rejection_reason': 'EAN validation failed'
                                                })

                                        # If product has NO accepted variants, treat as failed
                                        if not accepted_variants:
                                            failed_count += 1
                                            rejection_reasons = []
                                            for variant in rejected_variants:
                                                if hasattr(variant, 'rejection_reason'):
                                                    rejection_reasons.append(f"{variant.name}: {variant.rejection_reason}")
                                                else:
                                                    rejection_reasons.append(f"{variant.name}: Unknown rejection reason")
                                            
                                            error_message = 'All variants were rejected'
                                            if rejection_reasons:
                                                error_message += f' - {"; ".join(rejection_reasons)}'
                                            else:
                                                error_message += ' - Check variant data for validation errors'
                                            
                                            failed_products.append({
                                                'index': idx,
                                                'product_name': product.name,
                                                'product_sku': product.sku,
                                                'error': error_message,
                                                'rejected_variants_count': len(rejected_variants),
                                                'rejected_variants': [
                                                    {
                                                        'variant_name': v.name,
                                                        'variant_sku': v.sku,
                                                        'is_rejected': v.is_rejected
                                                    } for v in rejected_variants
                                                ],
                                                'original_data': item
                                            })
                                    else:
                                        failed_count += 1
                                        failed_products.append({
                                            'index': idx,
                                            'product_name': item.get('name', 'Unknown'),
                                            'error': 'Product creation failed - no product returned',
                                            'original_data': item
                                        })
                                else:
                                    # Handle direct product creation
                                    product = result
                                    product.created_by = request.user
                                    product.save()

                                    # Get all variants and separate accepted/rejected
                                    all_variants = product.variants.all()
                                    accepted_variants = [v for v in all_variants if not v.is_rejected]
                                    rejected_variants = [v for v in all_variants if v.is_rejected]

                                    print(f"Debug - Product {product.name}: total_variants={len(all_variants)}, accepted={len(accepted_variants)}, rejected={len(rejected_variants)}")
                                    for v in all_variants:
                                        print(f"  Variant {v.name}: is_rejected={v.is_rejected}")

                                    # Only include in created_products if product has at least one accepted variant
                                    if accepted_variants:
                                        created_products.append({
                                            'product_id': product.id,
                                            'product_name': product.name,
                                            'product_sku': product.sku,
                                            'total_variants': len(all_variants),
                                            'accepted_variants_count': len(accepted_variants),
                                            'rejected_variants_count': len(rejected_variants),
                                            'created_variants': [
                                                {
                                                    'variant_id': v.id,
                                                    'variant_name': v.name,
                                                    'variant_sku': v.sku,
                                                    'is_rejected': v.is_rejected
                                                }
                                                for v in accepted_variants  # Only show accepted variants
                                            ]
                                        })
                                        created_count += 1

                                    # If product has NO accepted variants, treat as failed
                                    if not accepted_variants:
                                        failed_count += 1
                                        rejection_reasons = []
                                        for variant in rejected_variants:
                                            if hasattr(variant, 'rejection_reason'):
                                                rejection_reasons.append(f"{variant.name}: {variant.rejection_reason}")
                                            else:
                                                rejection_reasons.append(f"{variant.name}: Unknown rejection reason")
                                        
                                        error_message = 'All variants were rejected'
                                        if rejection_reasons:
                                            error_message += f' - {"; ".join(rejection_reasons)}'
                                        else:
                                            error_message += ' - Check variant data for validation errors'
                                        
                                        failed_products.append({
                                            'index': idx,
                                            'product_name': product.name,
                                            'product_sku': product.sku,
                                            'error': error_message,
                                            'rejected_variants_count': len(rejected_variants),
                                            'rejected_variants': [
                                                {
                                                    'variant_name': v.name,
                                                    'variant_sku': v.sku,
                                                    'is_rejected': v.is_rejected
                                                } for v in rejected_variants
                                            ],
                                            'original_data': item
                                        })
                            else:
                                failed_count += 1
                                failed_products.append({
                                    'index': idx,
                                    'product_name': item.get('name', 'Unknown'),
                                    'error': f"Validation failed: {serializer.errors}",
                                    'validation_errors': serializer.errors,
                                    'original_data': item
                                })

                        except Exception as e:
                            failed_count += 1
                            failed_products.append({
                                'index': idx,
                                'product_name': item.get('name', 'Unknown'),
                                'error': f"Creation failed: {str(e)}",
                                'original_data': item
                            })

                    except Exception as e:
                        failed_count += 1
                        failed_products.append({
                            'index': idx,
                            'product_name': item.get('name', 'Unknown') if isinstance(item, dict) else 'Unknown',
                            'error': str(e),
                            'original_data': item
                        })

                print(f"⚡ Creation processing: {round((time.time() - process_start) * 1000, 1)}ms")

        except Exception as e:
            return Response({
                'error': f'Bulk create operation failed: {str(e)}',
                'total_products': len(request.data),
                'created_count': created_count if 'created_count' in locals() else 0,
                'failed_count': failed_count if 'failed_count' in locals() else 0,
                'failed_products': failed_products if 'failed_products' in locals() else []
            }, status=500)

        # Final performance summary
        total_time = time.time() - start_time
        print(f"🎯 BULK CREATE SUMMARY:")
        print(f"   ⚡ Total time: {round(total_time * 1000, 1)}ms")
        print(f"   📈 Created: {created_count} products")
        print(f"   ❌ Failed: {failed_count}")
        print(f"   🚫 EAN rejected: {len(ean_rejected_products)}")
        print(f"   🚀 Rate: {round(len(request.data) / total_time, 1)} products/second")

        return Response({
            "total_products": len(request.data),
            "created_count": created_count,
            "failed_count": failed_count,
            "created_products": created_products,
            "failed_products": failed_products,
            "ean_rejected_count": len(ean_rejected_products),
            "ean_rejected_products": ean_rejected_products,
            "api_time_seconds": round(total_time, 3),
            "performance_rate_per_second": round(len(request.data) / total_time, 1) if total_time > 0 else 0,
            "mode": "bulk_create",
            "message": f"Created {created_count} products in {round(total_time * 1000, 1)}ms"
        }, status=status.HTTP_201_CREATED)

class BulkUpdateProductsView(APIView):
    """
    PUT /api/products/bulk-update/
    Body: JSON array of product objects for UPDATING ONLY (SKUs required)
    All variants must have SKUs to identify which products/variants to update.

    Custom fields can be included in variant data under 'custom_fields' key in two formats:
    1. Array format: [{"field_id": 13, "value": "Yes"}]
    2. Dict format: {"field_name": "value"} (legacy)
    """
    permission_classes = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        from django.db import transaction
        import time

        start_time = time.time()
        print(f"🔄 BULK UPDATE MODE: Updating {len(request.data)} products...")

        try:
            with transaction.atomic():
                # Step 1: Validate that all variants have SKUs (update only)
                validation_start = time.time()
                all_skus_in_request = []

                for idx, item in enumerate(request.data):
                    if not isinstance(item, dict):
                        return Response({
                            'error': f'Item {idx + 1} must be a dictionary/object.',
                            'message': 'All items in the request must be valid JSON objects.'
                        }, status=400)

                    if 'variants' not in item or not isinstance(item['variants'], list):
                        return Response({
                            'error': f'Product {idx + 1} must have a "variants" array.',
                            'message': 'Each product must contain variants array for updating.'
                        }, status=400)

                    for variant_idx, variant_data in enumerate(item['variants']):
                        if not isinstance(variant_data, dict):
                            return Response({
                                'error': f'Variant {variant_idx + 1} in product {idx + 1} must be a dictionary.',
                                'message': 'All variants must be valid JSON objects.'
                            }, status=400)

                        if not variant_data.get('sku'):
                            return Response({
                                'error': f'SKU is required for variant {variant_idx + 1} in product {idx + 1}.',
                                'message': 'All variants must have SKUs for updating. Use bulk-create endpoint for new products.'
                            }, status=400)

                        all_skus_in_request.append(variant_data['sku'])

                # Check for duplicate SKUs in request and track them
                duplicate_skus = []
                seen_skus = set()
                duplicate_products = []
                requested_skus = set()

                for idx, item in enumerate(request.data):
                    item_duplicates = []
                    if 'variants' in item:
                        for variant_idx, variant_data in enumerate(item['variants']):
                            sku = variant_data.get('sku')
                            if sku:
                                if sku in seen_skus:
                                    duplicate_skus.append(sku)
                                    item_duplicates.append({
                                        'variant_index': variant_idx,
                                        'sku': sku,
                                        'variant_name': variant_data.get('name', f'Variant {variant_idx + 1}')
                                    })
                                else:
                                    seen_skus.add(sku)
                                    requested_skus.add(sku)

                    if item_duplicates:
                        duplicate_products.append({
                            'product_index': idx,
                            'product_name': item.get('name', f'Product {idx + 1}'),
                            'duplicate_variants': item_duplicates,
                            'reason': 'Duplicate SKUs found in request'
                        })

                print(f"⚡ SKU validation: {round((time.time() - validation_start) * 1000, 1)}ms - {len(requested_skus)} SKUs to update")

                # Step 2: Find existing variants
                lookup_start = time.time()
                existing_variants_by_sku = {
                    v.sku: v for v in ProductVariant.objects.select_related('product').filter(sku__in=requested_skus)
                }

                missing_skus = requested_skus - set(existing_variants_by_sku.keys())
                if missing_skus:
                    return Response({
                        'error': f'SKUs not found in database: {", ".join(list(missing_skus)[:10])}{"..." if len(missing_skus) > 10 else ""}',
                        'message': f'{len(missing_skus)} SKUs do not exist. Cannot update non-existent variants.',
                        'missing_skus': list(missing_skus)
                    }, status=400)

                print(f"⚡ Variant lookup: {round((time.time() - lookup_start) * 1000, 1)}ms - Found all {len(existing_variants_by_sku)} variants")

                # Step 3: EAN validation for variants that have EAN numbers
                ean_start = time.time()
                all_ean_numbers = []
                for item in request.data:
                    for variant_data in item['variants']:
                        if variant_data.get('ean_number'):
                            all_ean_numbers.append(variant_data['ean_number'])

                ean_validation_results = {}
                if all_ean_numbers and len(all_ean_numbers) <= 200:
                    try:
                        import os, requests, json
                        gs1_url = os.environ.get("GS1_API_URL")
                        gs1_token = os.environ.get("GS1_API_TOKEN", "")

                        if gs1_url and gs1_token:
                            session = requests.Session()
                            session.headers.update({'Authorization': f'Bearer {gs1_token}'})

                            gtin_param = json.dumps(all_ean_numbers)
                            response = session.get(
                                gs1_url,
                                params={'gtin': gtin_param, 'status': 'published'},
                                timeout=3
                            )

                            if response.status_code == 200:
                                data = response.json()
                                if data.get('status') and data.get('items'):
                                    for item_data in data['items']:
                                        ean = item_data.get('gtin')
                                        if ean:
                                            ean_validation_results[ean] = {
                                                'hsn_code': item_data.get('hs_code'),
                                                'tax': item_data.get('tax_rate', 0) or item_data.get('igst', 0),
                                                'cgst': item_data.get('cgst', 0),
                                                'sgst': item_data.get('sgst', 0),
                                                'igst': item_data.get('igst', 0),
                                                'cess': item_data.get('cess', 0),
                                                'is_valid': True
                                            }
                            session.close()
                    except Exception:
                        pass

                # Mark unvalidated EANs
                for ean in all_ean_numbers:
                    if ean not in ean_validation_results:
                        ean_validation_results[ean] = {'is_valid': False}

                print(f"⚡ EAN validation: {round((time.time() - ean_start) * 1000, 1)}ms - {len(ean_validation_results)} EANs checked")

                # Step 4: Update products and variants
                update_start = time.time()
                updated_products_list = []
                updated_products_set = set()
                updated_variants_count = 0
                failed_updates = []

                for idx, item in enumerate(request.data):
                    try:
                        # Check if this product has duplicates and skip if so
                        has_duplicates = any(dp['product_index'] == idx for dp in duplicate_products)
                        if has_duplicates:
                            continue

                        # Group variants by product (first variant determines the product)
                        first_variant_sku = item['variants'][0]['sku']

                        # Check if first variant SKU exists (skip missing SKUs)
                        if first_variant_sku not in existing_variants_by_sku:
                            failed_updates.append({
                                'product_index': idx,
                                'product_name': item.get('name', 'Unknown'),
                                'error': f'First variant SKU "{first_variant_sku}" not found in database',
                                'original_data': item
                            })
                            continue

                        existing_product = existing_variants_by_sku[first_variant_sku].product

                        # Update product fields if provided
                        if 'name' in item:
                            existing_product.name = item['name']
                        if 'description' in item:
                            existing_product.description = item['description']
                        if 'category' in item:
                            existing_product.category_id = item['category']
                        if 'brand' in item:
                            existing_product.brand_id = item['brand']
                        if 'is_active' in item:
                            existing_product.is_active = item['is_active']
                        if 'is_published' in item:
                            existing_product.is_published = item['is_published']
                        if 'tags' in item:
                            existing_product.tags = item['tags']

                        existing_product.updated_by = request.user
                        existing_product.save()
                        updated_products_set.add(existing_product.id)

                        # Track updated variants for this product
                        updated_variants = []

                        # Update variants
                        for variant_data in item['variants']:
                            sku = variant_data['sku']

                            # Skip if this variant's SKU is not in the database
                            if sku not in existing_variants_by_sku:
                                failed_updates.append({
                                    'product_index': idx,
                                    'product_name': item.get('name', 'Unknown'),
                                    'error': f'Variant SKU "{sku}" not found in database',
                                    'sku': sku,
                                    'original_data': item
                                })
                                continue

                            existing_variant = existing_variants_by_sku[sku]

                            # Apply EAN validation if provided
                            ean_number = variant_data.get('ean_number')
                            if ean_number and ean_number in ean_validation_results:
                                validation = ean_validation_results[ean_number]
                                if validation.get('is_valid', True):
                                    existing_variant.hsn_code = validation.get('hsn_code', existing_variant.hsn_code)
                                    existing_variant.tax = validation.get('tax', existing_variant.tax)
                                    existing_variant.cgst = validation.get('cgst', existing_variant.cgst)
                                    existing_variant.sgst = validation.get('sgst', existing_variant.sgst)
                                    existing_variant.igst = validation.get('igst', existing_variant.igst)
                                    existing_variant.cess = validation.get('cess', existing_variant.cess)

                            # Update all provided variant fields
                            field_mapping = {
                                'name': 'name',
                                'description': 'description',
                                'tags': 'tags',
                                'base_price': 'base_price',
                                'mrp': 'mrp',
                                'selling_price': 'selling_price',
                                'ean_number': 'ean_number',
                                'ran_number': 'ran_number',
                                'hsn_code': 'hsn_code',
                                'tax': 'tax',
                                'cgst': 'cgst',
                                'sgst': 'sgst',
                                'igst': 'igst',
                                'cess': 'cess',
                                'weight': 'weight',
                                'net_qty': 'net_qty',
                                'packaging_type': 'packaging_type',
                                'product_dimensions': 'product_dimensions',
                                'package_dimensions': 'package_dimensions',
                                'shelf_life': 'shelf_life',
                                'uom': 'uom',
                                'attributes': 'attributes',
                                'is_pack': 'is_pack',
                                'pack_qty': 'pack_qty',
                                'is_active': 'is_active',
                                'is_b2b_enable': 'is_b2b_enable',
                                'is_pp_enable': 'is_pp_enable',
                                'is_visible': 'is_visible',
                                'is_published': 'is_published',
                                'is_rejected': 'is_rejected'
                            }

                            for field_name, model_field in field_mapping.items():
                                if field_name in variant_data:
                                    setattr(existing_variant, model_field, variant_data[field_name])

                            existing_variant.save()
                            updated_variants_count += 1

                            # Track this updated variant
                            updated_variants.append({
                                'variant_id': existing_variant.id,
                                'sku': existing_variant.sku,
                                'name': existing_variant.name,
                                'base_price': float(existing_variant.base_price) if existing_variant.base_price else 0,
                                'mrp': float(existing_variant.mrp) if existing_variant.mrp else 0,
                                'selling_price': float(existing_variant.selling_price) if existing_variant.selling_price else 0
                            })

                            # Handle variant images if provided
                            if 'images' in variant_data:
                                self._handle_variant_images(existing_variant, variant_data['images'])

                            # Handle custom fields if provided
                            if 'custom_fields' in variant_data:
                                self._handle_variant_custom_fields(existing_variant, variant_data['custom_fields'])

                        # Add this product to the updated products list
                        updated_products_list.append({
                            'product_id': existing_product.id,
                            'product_name': existing_product.name,
                            'sku': existing_product.sku,
                            'updated_variants_count': len(updated_variants),
                            'updated_variants': updated_variants
                        })

                    except Exception as e:
                        failed_updates.append({
                            'index': idx,
                            'product_name': item.get('name', 'Unknown'),
                            'error': str(e),
                            'original_data': item
                        })

                print(f"⚡ Update processing: {round((time.time() - update_start) * 1000, 1)}ms")

                total_time = time.time() - start_time
                print(f"🎯 BULK UPDATE SUMMARY:")
                print(f"   ⚡ Total time: {round(total_time * 1000, 1)}ms")
                print(f"   🔄 Updated products: {len(updated_products_set)}")
                print(f"   🔄 Updated variants: {updated_variants_count}")
                print(f"   ❌ Failed updates: {len(failed_updates)}")
                print(f"   🔀 Duplicate products: {len(duplicate_products)}")

                return Response({
                    "total_products": len(request.data),
                    "updated_products_count": len(updated_products_set),
                    "updated_variants_count": updated_variants_count,
                    "duplicate_products_count": len(duplicate_products),
                    "failed_updates_count": len(failed_updates),
                    "updated_products": updated_products_list,
                    "duplicate_products": duplicate_products,
                    "failed_updates": failed_updates,
                    "duplicate_skus": list(set(duplicate_skus)),
                    "api_time_seconds": round(total_time, 3),
                    "performance_rate_per_second": round(len(request.data) / total_time, 1) if total_time > 0 else 0,
                    "mode": "bulk_update",
                    "message": f"Updated {len(updated_products_set)} products with {updated_variants_count} variants in {round(total_time * 1000, 1)}ms"
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'error': f'Bulk update operation failed: {str(e)}',
                'message': 'An unexpected error occurred during the bulk update process.'
            }, status=500)


class ProductExportView(APIView):
    permission_classes = [IsAuthenticated]
    """
    GET /api/products/export/
    Exports every Product×Variant as a row in Excel, showing names for category, brand, and collections.
    """

    def get(self, request, *args, **kwargs):

        # Lightweight validation - check required fields only
        for idx, item in enumerate(request.data):
            try:
                if not isinstance(item, dict):
                    errors[idx] = ["Item must be a dictionary"]
                    continue

                # Check only essential fields
                if not item.get('name'):
                    errors[idx] = ["Name is required"]
                    continue

                if not item.get('category'):
                    errors[idx] = ["Category is required"]
                    continue

                variants = item.get('variants', [])
                if not variants:
                    errors[idx] = ["At least one variant is required"]
                    continue

                # Quick variant validation
                valid_variants = []
                for variant in variants:
                    if isinstance(variant, dict) and variant.get('sku'):
                        valid_variants.append(variant)

                if not valid_variants:
                    errors[idx] = ["At least one variant with SKU is required"]
                    continue

                item['variants'] = valid_variants
                valid_items.append(item)

            except Exception as e:
                errors[idx] = [str(e)]

        print(f"📋 Lightweight validation completed in {round(time.time() - validation_start, 2)}s - {len(valid_items)}/{len(request.data)} items valid")

        # OPTIMIZATION 2: Batch EAN validation for all products at once
        ean_validation_results = {}
        all_ean_numbers = []

        # Collect all EAN numbers from valid items (check for new/updated variants by SKU)
        for item in valid_items:
            variants_data = item.get('variants', [])
            for variant_data in variants_data:
                sku = variant_data.get('sku')
                ean_number = variant_data.get('ean_number')
                # Validate EAN for new variants or variants with different EAN than existing
                if ean_number and (not sku or sku not in existing_variants_by_sku):
                    all_ean_numbers.append(ean_number)

        # OPTIMIZATION: Skip EAN validation for bulk operations to maximize speed
        # EAN validation can be done separately via a background job if needed
        ean_start = time.time()
        print(f"⚡ Skipping EAN validation for bulk operation speed (found {len(all_ean_numbers)} EANs)")

        # Mark all EANs as valid to allow creation
        for ean in all_ean_numbers:
            ean_validation_results[ean] = {'is_valid': True}

        print(f"✅ EAN validation skipped in {round(time.time() - ean_start, 3)}s")

        # OPTIMIZATION 3: Separate items into create vs update batches for bulk operations
        items_to_create = []
        items_to_update = []
        failed_products = []
        ean_rejected_products = []

        # Pre-sort items into create/update categories
        for idx, item in enumerate(valid_items):
            try:
                variants_data = item.get('variants', [])
                if not variants_data:
                    failed_products.append({
                        'index': idx,
                        'product_name': item.get('name', 'Unknown'),
                        'error': 'No variants provided'
                    })
                    continue

                # Check if any variant SKU already exists
                existing_variant = None
                for variant_data in variants_data:
                    sku = variant_data.get('sku')
                    if sku and sku in existing_variants_by_sku:
                        existing_variant = existing_variants_by_sku[sku]
                        break

                if existing_variant:
                    items_to_update.append((idx, item, existing_variant))
                else:
                    items_to_create.append((idx, item))

            except Exception as e:
                failed_products.append({
                    'index': idx,
                    'product_name': item.get('name', 'Unknown'),
                    'error': str(e)
                })

        # BULK CREATE OPERATIONS
        created_products = []
        products_to_create = []
        variants_to_create = []

        for idx, item in items_to_create:
            try:
                # Prepare product data
                clean_item = item.copy()
                clean_item.pop('id', None)
                variants_data = clean_item.pop('variants', [])

                # Create product instance (not saved yet)
                product = Product(
                    name=clean_item['name'],
                    description=clean_item.get('description', ''),
                    tags=clean_item.get('tags', []),
                    category=clean_item['category'],
                    brand=clean_item.get('brand'),
                    is_active=clean_item.get('is_active', True),
                    is_published=clean_item.get('is_published', True),
                    image=clean_item.get('image', ''),
                    created_by=request.user,
                    updated_by=request.user,
                )
                products_to_create.append((product, variants_data, idx))

            except Exception as e:
                failed_products.append({
                    'index': idx,
                    'product_name': item.get('name', 'Unknown'),
                    'error': str(e)
                })

        # SUPER OPTIMIZED: Bulk create ALL variants in one operation
        bulk_create_start = time.time()
        all_variants_to_create = []

        if products_to_create:
            print(f"🏭 Starting super-fast bulk creation of {len(products_to_create)} products...")
            try:
                # Extract just the product objects for bulk_create
                product_objects = [product for product, _, _ in products_to_create]
                Product.objects.bulk_create(product_objects)

                # Prepare ALL variants for bulk creation
                variants_with_custom_fields = []  # Store custom fields separately
                for product, variants_data, idx in products_to_create:
                    for variant_data in variants_data:
                        # Clean variant data
                        clean_variant = {
                            'product': product,
                            'name': variant_data.get('name', ''),
                            'sku': variant_data.get('sku', ''),
                            'ean_number': variant_data.get('ean_number', ''),
                            'net_qty': variant_data.get('net_qty', ''),
                            'base_price': variant_data.get('base_price', 0),
                            'mrp': variant_data.get('mrp', 0),
                            'selling_price': variant_data.get('selling_price', 0),
                            'weight': variant_data.get('weight', 0),
                            'is_active': variant_data.get('is_active', True),
                            'color': variant_data.get('color', ''),
                            'size': variant_data.get('size', ''),
                            'is_rejected': False  # Skip EAN validation for speed
                        }
                        variant_obj = ProductVariant(**clean_variant)
                        all_variants_to_create.append(variant_obj)

                        # Store custom fields for later processing
                        custom_fields_data = variant_data.get('custom_fields', [])
                        if custom_fields_data:
                            variants_with_custom_fields.append({
                                'sku': variant_data.get('sku'),
                                'custom_fields': custom_fields_data
                            })

                    created_products.append(product)

                # Bulk create ALL variants at once (massive performance gain)
                if all_variants_to_create:
                    ProductVariant.objects.bulk_create(all_variants_to_create, batch_size=1000)

                    # Handle custom fields for bulk created variants
                    if variants_with_custom_fields:
                        from cms.models.product import ProductVariantCustomField
                        from cms.models.setting import CustomField
                        custom_field_values = []

                        for variant_custom_data in variants_with_custom_fields:
                            sku = variant_custom_data['sku']
                            custom_fields_data = variant_custom_data['custom_fields']

                            # Get the most recently created variant with this SKU to avoid duplicates
                            created_variant = ProductVariant.objects.filter(sku=sku).order_by('-id').first()
                            if not created_variant:
                                continue

                            # Support both dict format and array format
                            if isinstance(custom_fields_data, dict):
                                # Legacy dict format: {"field_name": "value"}
                                for field_name, field_value in custom_fields_data.items():
                                    try:
                                        custom_field = CustomField.objects.get(name=field_name, is_active=True)
                                        custom_field_values.append(
                                            ProductVariantCustomField(
                                                product_variant=created_variant,
                                                custom_field=custom_field,
                                                value=str(field_value)
                                            )
                                        )
                                    except CustomField.DoesNotExist:
                                        continue
                            elif isinstance(custom_fields_data, list):
                                # New array format: [{"field_id": 13, "value": "Yes"}]
                                for field_item in custom_fields_data:
                                    if isinstance(field_item, dict) and 'field_id' in field_item:
                                        try:
                                            custom_field = CustomField.objects.get(id=field_item['field_id'], is_active=True)
                                            custom_field_values.append(
                                                ProductVariantCustomField(
                                                    product_variant=created_variant,
                                                    custom_field=custom_field,
                                                    value=str(field_item.get('value', ''))
                                                )
                                            )
                                        except CustomField.DoesNotExist:
                                            continue

                        # Bulk create custom field values
                        if custom_field_values:
                            ProductVariantCustomField.objects.bulk_create(
                                custom_field_values,
                                batch_size=1000,
                                ignore_conflicts=True
                            )

                print(f"✅ Super-fast bulk creation completed in {round(time.time() - bulk_create_start, 2)}s")
                print(f"   📦 Created {len(created_products)} products with {len(all_variants_to_create)} variants")

            except Exception as e:
                print(f"Bulk create error: {e}")
                # Fallback to product-by-product creation
                created_products = []
                for product, variants_data, idx in products_to_create:
                    try:
                        product.save()
                        for variant_data in variants_data:
                            clean_variant = {
                                'name': variant_data.get('name', ''),
                                'sku': variant_data.get('sku', ''),
                                'ean_number': variant_data.get('ean_number', ''),
                                'net_qty': variant_data.get('net_qty', ''),
                                'base_price': variant_data.get('base_price', 0),
                                'mrp': variant_data.get('mrp', 0),
                                'selling_price': variant_data.get('selling_price', 0),
                                'weight': variant_data.get('weight', 0),
                                'is_active': variant_data.get('is_active', True),
                                'color': variant_data.get('color', ''),
                                'size': variant_data.get('size', ''),
                                'is_rejected': False
                            }
                            created_variant = ProductVariant.objects.create(product=product, **clean_variant)

                            # Handle custom fields for fallback variant creation
                            custom_fields_data = variant_data.get('custom_fields', [])
                            if custom_fields_data:
                                from cms.models.product import ProductVariantCustomField
                                from cms.models.setting import CustomField

                                # Support both dict format and array format
                                if isinstance(custom_fields_data, dict):
                                    # Legacy dict format: {"field_name": "value"}
                                    for field_name, field_value in custom_fields_data.items():
                                        try:
                                            custom_field = CustomField.objects.get(name=field_name, is_active=True)
                                            ProductVariantCustomField.objects.create(
                                                product_variant=created_variant,
                                                custom_field=custom_field,
                                                value=str(field_value)
                                            )
                                        except CustomField.DoesNotExist:
                                            continue
                                elif isinstance(custom_fields_data, list):
                                    # New array format: [{"field_id": 13, "value": "Yes"}]
                                    for field_item in custom_fields_data:
                                        if isinstance(field_item, dict) and 'field_id' in field_item:
                                            try:
                                                custom_field = CustomField.objects.get(id=field_item['field_id'], is_active=True)
                                                ProductVariantCustomField.objects.create(
                                                    product_variant=created_variant,
                                                    custom_field=custom_field,
                                                    value=str(field_item.get('value', ''))
                                                )
                                            except CustomField.DoesNotExist:
                                                continue
                        created_products.append(product)
                    except Exception as individual_error:
                        failed_products.append({
                            'index': idx,
                            'product_name': product.name,
                            'error': str(individual_error)
                        })

        # BULK UPDATE OPERATIONS (keep individual for now due to complexity)
        updated_products = []
        for idx, item, existing_variant in items_to_update:
            try:
                product = existing_variant.product

                # Update product fields
                product.name = item['name']
                product.description = item.get('description', '')
                product.tags = item.get('tags', [])
                product.category = item['category']
                product.brand = item.get('brand')
                product.is_active = item.get('is_active', True)
                product.is_published = item.get('is_published', True)
                product.image = item.get('image', '')
                product.updated_by = request.user
                product.save()

                # Update variants (could be optimized further with bulk_update in future)
                variants_data = item.get('variants', [])
                for variant_data in variants_data:
                    variant_data = variant_data.copy()
                    variant_data.pop('id', None)
                    variant_images = variant_data.pop('images', [])
                    variant_data.pop('link', None)

                    sku = variant_data.get('sku')
                    if sku and sku in existing_variants_by_sku:
                        # Update existing variant
                        variant = existing_variants_by_sku[sku]
                        for field, value in variant_data.items():
                            setattr(variant, field, value)
                        variant.save()

                        # Handle images for existing variant
                        if variant_images:
                            self._handle_variant_images(variant, variant_images)
                    else:
                        # Create new variant
                        new_variant = ProductVariant.objects.create(product=product, **variant_data)

                        # Handle images for new variant
                        if variant_images:
                            self._handle_variant_images(new_variant, variant_images)

                updated_products.append(product)

            except Exception as e:
                failed_products.append({
                    'index': idx,
                    'product_name': item.get('name', 'Unknown'),
                    'error': str(e)
                })

        end_time = time.time()
        api_time = round(end_time - start_time, 2)

        # Performance summary
        print(f"🎯 PERFORMANCE SUMMARY:")
        print(f"   📊 Total time: {api_time}s")
        print(f"   📈 Created: {len(created_products)} products")
        print(f"   🔄 Updated: {len(updated_products)} products")
        print(f"   ❌ Failed: {len(failed_products)} products")
        print(f"   ⚡ Rate: {round(len(request.data) / api_time, 1)} products/second")

        status_code = status.HTTP_201_CREATED if not errors and not failed_products else status.HTTP_207_MULTI_STATUS
        return Response(
            {
                "total_products": len(request.data),
                "total_variants": sum(len(item.get('variants', [])) for item in request.data),
                "created_count": len(created_products),
                "updated_count": len(updated_products),
                "failed_count": len(failed_products),
                "failed_products": failed_products,
                "ean_rejected_count": len(ean_rejected_products),
                "ean_rejected_products": ean_rejected_products,
                "validation_errors_count": len(errors),
                "validation_errors": errors,
                "api_time_seconds": api_time,
                "performance_rate_per_second": round(len(request.data) / api_time, 2) if api_time > 0 else 0
            },
            status=status_code
        )
    
class ProductExportView(APIView):
    permission_classes = [AllowAny]
    """
    GET /api/products/export/
    Exports products data in Excel format with enhanced formatting and comprehensive data.

    Query Parameters:
    - status: Filter by product status (active/inactive)
    - category: Filter by category ID
    - brand: Filter by brand ID
    - collection: Filter by collection ID
    - search: Search in product name, description, tags, SKU
    - ordering: Order by field (name, category__name, brand__name, is_active, created_date, updation_date) - aligned with ProductViewSet
    - order_direction: Order direction (asc/desc) - default: asc
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum number of records to export (default: 1000, max: 10000)
    - include_images: Include image data (true/false) - default: false
    - include_custom_fields: Include custom fields (true/false) - default: true
    - include_size_chart: Include size chart data (true/false) - default: true
    """

    def get(self, request, *args, **kwargs):
        import datetime
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Get query parameters (aligned with ProductViewSet)
        is_active = request.GET.get('status')
        category = request.GET.get('category')
        brand = request.GET.get('brand')
        collection = request.GET.get('collection')
        search_query = request.GET.get('search', '')

        # Ordering parameters (matching ProductViewSet)
        ordering = request.GET.get('ordering', 'updation_date')  # Default from ProductViewSet
        order_direction = request.GET.get('order_direction', 'asc').lower()

        # Export options
        export_format = request.GET.get('format', 'excel').lower()
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)  # Default 1000, max 10000 to prevent timeouts
        include_images = request.GET.get('include_images', 'false').lower() == 'true'
        include_custom_fields = request.GET.get('include_custom_fields', 'true').lower() == 'true'
        include_size_chart = request.GET.get('include_size_chart', 'true').lower() == 'true'

        # Build queryset with optimized prefetching
        qs = ProductVariant.objects.select_related(
            'product', 'product__category', 'product__brand'
        ).prefetch_related(
            'product__collections', 'images', 'custom_field_values__custom_field', 'size_chart_values'
        )

        # Apply filters
        if is_active is not None:
            qs = qs.filter(product__is_active=is_active.lower() == 'true')
        if category:
            qs = qs.filter(product__category_id=category)
        if brand:
            qs = qs.filter(product__brand_id=brand)
        if collection:
            qs = qs.filter(product__collections__id=collection)
        if search_query:
            search_filter = (
                Q(product__name__icontains=search_query) |
                Q(product__description__icontains=search_query) |
                Q(product__tags__icontains=search_query) |
                Q(sku__icontains=search_query)
            )
            qs = qs.filter(search_filter)

        # Apply ordering (matching ProductViewSet ordering_fields)
        # ProductViewSet ordering_fields: ['name', 'category__name', 'brand__name', 'is_active', 'created_date', 'updation_date']
        order_field = self._get_export_order_field(ordering)
        if order_direction == 'desc':
            order_field = f'-{order_field}'

        # Always add secondary ordering for consistency
        qs = qs.order_by(order_field, 'product__name', 'name')

        # Apply limit to prevent timeouts
        qs = qs[:export_limit]

        # Serialize data
        data = ProductExportSerializer(qs, many=True).data

        # Filter fields based on parameters
        if data and not include_custom_fields:
            for item in data:
                item.pop('variant_custom_fields', None)

        if data and not include_size_chart:
            for item in data:
                item.pop('variant_size_chart', None)

        if data and not include_images:
            for item in data:
                item.pop('variant_images_count', None)
                item.pop('variant_primary_image', None)

        # Handle different export formats
        if export_format == 'csv':
            return self._export_csv(data, request)
        else:
            return self._export_excel(data, request)

    def _export_csv(self, data, request):
        """Export data as CSV with custom fields as separate columns"""
        import csv
        import datetime
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="products_export_{timestamp}.csv"'

        if data:
            # Get all unique custom field columns
            custom_field_columns = self._get_custom_field_columns(data)

            # Build headers (excluding variant_custom_fields)
            base_headers = [h for h in data[0].keys() if h != 'variant_custom_fields']

            # Insert custom field columns after variant_attributes
            insert_index = base_headers.index('variant_attributes') + 1 if 'variant_attributes' in base_headers else len(base_headers)
            for i, cf_name in enumerate(custom_field_columns):
                base_headers.insert(insert_index + i, cf_name)

            writer = csv.DictWriter(response, fieldnames=base_headers)
            writer.writeheader()

            # Write data rows with custom fields expanded
            for item in data:
                # Parse custom fields for this item
                custom_fields_dict = self._parse_custom_fields(item.get('variant_custom_fields', ''))

                # Build row with expanded custom fields
                row = {}
                for header in base_headers:
                    if header in custom_field_columns:
                        row[header] = custom_fields_dict.get(header, '')
                    else:
                        row[header] = item.get(header, '')

                writer.writerow(row)

        return response

    def _export_excel(self, data, request):
        """Export data as Excel with enhanced formatting and custom fields as separate columns"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products Export"

        # Get all unique custom fields across all categories in the export
        custom_field_columns = self._get_custom_field_columns(data)

        # Get base headers (excluding variant_custom_fields)
        if data:
            headers = [h for h in list(data[0].keys()) if h != 'variant_custom_fields']

            # Insert custom field columns after variant_attributes
            insert_index = headers.index('variant_attributes') + 1 if 'variant_attributes' in headers else len(headers)
            for i, cf_name in enumerate(custom_field_columns):
                headers.insert(insert_index + i, cf_name)
        else:
            headers = ["product_id", "product_title", "variant_title", "variant_sku", "product_category", "product_brand"]

        # Format headers for display
        def format_header(header):
            return header.replace('_', ' ').title()

        formatted_headers = [format_header(h) for h in headers]

        # Add metadata row
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([f"Products Export - Generated on {timestamp}"])
        ws.append([f"Total Records: {len(data)}"])
        ws.append([])  # Empty row

        # Add headers with formatting
        header_row = ws.max_row + 1
        for col_num, (header, original_header) in enumerate(zip(formatted_headers, headers), 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")

            # Use different color for custom field columns
            if original_header in custom_field_columns:
                # Green background for custom fields
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            else:
                # Default blue background
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        # Add data rows with custom fields
        for item in data:
            row_data = []
            # Parse custom fields for this item
            custom_fields_dict = self._parse_custom_fields(item.get('variant_custom_fields', ''))

            for header in headers:
                # Check if this is a custom field column
                if header in custom_field_columns:
                    value = custom_fields_dict.get(header, '')
                else:
                    value = item.get(header, '')
                    # Format boolean values
                    if isinstance(value, bool):
                        value = 'Yes' if value else 'No'
                row_data.append(value)
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set column width with limits
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add filters to header row
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

        # Style data rows with alternating colors
        for row_num in range(header_row + 1, ws.max_row + 1):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )

        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response["Content-Disposition"] = f'attachment; filename="products_export_{timestamp}.xlsx"'

        wb.save(response)
        return response

    def _get_export_order_field(self, ordering):
        """
        Map ordering parameter to actual field name for ProductVariant queryset.
        Aligned with ProductViewSet ordering_fields: ['name', 'category__name', 'brand__name', 'is_active', 'created_date', 'updation_date']
        """
        order_mapping = {
            # ProductViewSet compatible fields (mapped to ProductVariant fields)
            'name': 'product__name',
            'category__name': 'product__category__name',
            'brand__name': 'product__brand__name',
            'is_active': 'product__is_active',
            'created_date': 'product__creation_date',
            'updation_date': 'product__updation_date',

            # Additional alias mappings for convenience
            'category': 'product__category__name',
            'brand': 'product__brand__name',
            'status': 'product__is_active',
            'created': 'product__creation_date',
            'updated': 'product__updation_date',

            # Variant-specific fields
            'variant_name': 'name',
            'sku': 'sku',
            'price': 'selling_price',
            'base_price': 'base_price',
            'mrp': 'mrp'
        }
        return order_mapping.get(ordering, 'product__updation_date')

    def _get_custom_field_columns(self, data):
        """
        Extract all unique custom field names from the data.
        Returns a list of unique custom field names to use as column headers.
        """
        custom_fields_set = set()

        for item in data:
            custom_fields_str = item.get('variant_custom_fields', '')
            if custom_fields_str:
                # Parse the custom fields string
                # Format: "Field1: Value1; Field2: Value2"
                fields_dict = self._parse_custom_fields(custom_fields_str)
                custom_fields_set.update(fields_dict.keys())

        # Return sorted list for consistent column order
        return sorted(list(custom_fields_set))

    def _parse_custom_fields(self, custom_fields_str):
        """
        Parse custom fields string into a dictionary.
        Input format: "Field Name: Value; Another Field: Another Value"
        Output: {"Field Name": "Value", "Another Field": "Another Value"}
        """
        fields_dict = {}

        if not custom_fields_str:
            return fields_dict

        # Split by semicolon to get individual field-value pairs
        pairs = custom_fields_str.split(';')

        for pair in pairs:
            pair = pair.strip()
            if ':' in pair:
                # Split by first colon only
                key, value = pair.split(':', 1)
                fields_dict[key.strip()] = value.strip()

        return fields_dict


class ProductPricingViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Unified API endpoint that returns all products with their pricing information.
    Can show either cluster pricing, facility pricing, or both based on query parameters.
    
    Query Parameters:
    - type: 'cluster' for cluster pricing, 'facility' for facility pricing
    - category: Filter by category ID
    - brand: Filter by brand ID
    - name: Filter by product name (partial match)
    - status: Filter by is_active (true/false)
    - is_published: Filter by is_published (true/false)
    - min_price: Filter by minimum price
    - max_price: Filter by maximum price
    """
    queryset = Product.objects.filter(is_active=True, is_published=True)
    permission_classes = [AllowAny]  # Adjust permissions as needed
    pagination_class = CustomPageNumberPagination
    
    def get_serializer_class(self):
        """Return appropriate serializer based on type parameter"""
        pricing_type = self.request.query_params.get('type', 'cluster')
        
        if pricing_type == 'facility':
            return ProductWithFacilityPricingSerializer
        else:  # default to cluster
            return ProductWithClusterPricingSerializer
    
    def get_queryset(self):
        """Override to add any additional filtering if needed"""
        queryset = super().get_queryset()
        
        # Add any additional filtering here if needed
        # For example, filter by category, brand, etc.
        category_id = self.request.query_params.get('category', None)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
            
        brand_id = self.request.query_params.get('brand', None)
        if brand_id:
            queryset = queryset.filter(brand_id=brand_id)
        
        # Filter by product name (case-insensitive partial match)
        product_name = self.request.query_params.get('name', None)
        if product_name:
            queryset = queryset.filter(name__icontains=product_name)
        
        # Add status filtering
        status = self.request.query_params.get('status', None)
        if status is not None:
            queryset = queryset.filter(is_active=status.lower() == 'true')
        
        is_published = self.request.query_params.get('is_published', None)
        if is_published is not None:
            queryset = queryset.filter(is_published=is_published.lower() == 'true')
            
        # Filter by price range
        min_price = self.request.query_params.get('min_price', None)
        if min_price:
            try:
                queryset = queryset.filter(variants__base_price__gte=float(min_price))
            except ValueError:
                pass
                
        max_price = self.request.query_params.get('max_price', None)
        if max_price:
            try:
                queryset = queryset.filter(variants__base_price__lte=float(max_price))
            except ValueError:
                pass
            
        return queryset.select_related('category', 'brand').prefetch_related('variants')


class ProductClusterPriceUpdateView(APIView):
    """
    API endpoint to update product pricing for a specific cluster.
    Updates both price and csp for all variants in all facilities of the cluster.
    """
    permission_classes = [AllowAny]  # Adjust permissions as needed
    
    def put(self, request, product_id):
        try:
            # Get the product
            product = Product.objects.get(id=product_id, is_active=True, is_published=True)
        except Product.DoesNotExist:
            return Response(
                {"error": "Product not found or inactive"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validate request data
        serializer = ClusterPriceUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        cluster_id = serializer.validated_data['cluster_id']
        margin = serializer.validated_data['margin']
        
        try:
            # Get the cluster
            from cms.models.facility import Cluster
            cluster = Cluster.objects.get(id=cluster_id, is_active=True)
        except Cluster.DoesNotExist:
            return Response(
                {"error": "Cluster not found or inactive"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get all facilities in the cluster
        cluster_facilities = cluster.facilities.all()
        if not cluster_facilities.exists():
            return Response(
                {"error": f"No facilities found in cluster '{cluster.name}'"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all active variants for the product
        product_variants = product.variants.filter(is_active=True)
        if not product_variants.exists():
            return Response(
                {"error": "No active variants found for this product"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update pricing for all variants in all facilities of the cluster
        updated_records = []
        total_updated = 0
        price_history_records = []
        
        # Get the current user (if authenticated)
        current_user = request.user if hasattr(request, 'user') and request.user.is_authenticated else None
        
        for facility in cluster_facilities:
            for variant in product_variants:
                # Get or create FacilityInventory
                inventory, created = FacilityInventory.objects.get_or_create(
                    facility=facility,
                    product_variant=variant,
                    defaults={'base_price': 0.0, 'selling_price': 0.0, 'is_active': True}
                )
                
                # Calculate new prices based on current selling price
                current_selling_price = inventory.selling_price if inventory.selling_price and inventory.selling_price > 0 else (variant.base_price or 0.0)
                if current_selling_price and current_selling_price > 0:  # Only update if current selling price exists
                    new_price = current_selling_price * (1 + margin / 100)
                    new_csp = current_selling_price * (1 + margin / 100)
                    
                    # Check if new price exceeds MRP
                    variant_mrp = variant.mrp or 0.0
                    if variant_mrp > 0 and new_price > variant_mrp:
                        return Response({
                            "error": f"Price update failed: New price {new_price:.2f} exceeds MRP {variant_mrp:.2f} for variant '{variant.name}' (Product: {product.name})",
                            "details": {
                                "product_id": product.id,
                                "product_name": product.name,
                                "variant_id": variant.id,
                                "variant_name": variant.name,
                                "base_price": current_selling_price,
                                "calculated_price": new_price,
                                "mrp": variant_mrp,
                                "margin": margin
                            }
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Store old values for response and history
                    old_price = inventory.selling_price if inventory.selling_price > 0 else current_selling_price
                    old_csp = inventory.selling_price if inventory.selling_price > 0 else current_selling_price
                    
                    # Update only selling price
                    inventory.selling_price = new_price
                    inventory.save()
                    
                    # Create price history record
                    price_history = ProductPriceHistory.objects.create(
                        product=product,
                        product_variant=variant,
                        cluster=cluster,
                        facility=facility,
                        user=current_user,
                        old_price=old_price,
                        new_price=new_price,
                        old_csp=old_csp,
                        new_csp=new_csp,
                        percentage_change=margin,
                        change_type='percentage_update',
                        change_reason=f'Price updated by {margin}% for cluster {cluster.name}'
                    )
                    price_history_records.append(price_history)
                    
                    updated_records.append({
                        'variant_id': variant.id,
                        'variant_name': variant.name,
                        'facility_id': facility.id,
                        'facility_name': facility.name,
                        'old_price': old_price,
                        'new_price': new_price,
                        'old_csp': old_csp,
                        'new_csp': new_csp,
                        'history_id': price_history.id
                    })
                    total_updated += 1
        
        if total_updated == 0:
            return Response(
                {"error": "No records updated. Product variants may not have base prices set."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            "success": True,
            "message": f"Prices updated successfully for cluster '{cluster.name}'",
            "product_id": product_id,
            "product_name": product.name,
            "cluster_id": cluster_id,
            "cluster_name": cluster.name,
            "margin": margin,
            "updated_records": total_updated,
            "history_records_created": len(price_history_records),
            "updated_pricing": updated_records,
            "price_history_ids": [record.id for record in price_history_records]
        }, status=status.HTTP_200_OK)


class ProductPriceHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint to view product price history
    """
    serializer_class = ProductPriceHistorySerializer
    permission_classes = [AllowAny]  # Adjust permissions as needed
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        """Filter price history based on query parameters"""
        queryset = ProductPriceHistory.objects.all().select_related(
            'product', 'product_variant', 'cluster', 'facility', 'user'
        )
        
        # Filter by product ID
        product_id = self.request.query_params.get('product_id')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        # Filter by cluster ID
        cluster_id = self.request.query_params.get('cluster_id')
        if cluster_id:
            queryset = queryset.filter(cluster_id=cluster_id)
        
        # Filter by facility ID
        facility_id = self.request.query_params.get('facility_id')
        if facility_id:
            queryset = queryset.filter(facility_id=facility_id)
        
        # Filter by user ID
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(creation_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(creation_date__lte=end_date)
        
        return queryset.order_by('-creation_date')


# class BulkPriceUpdateView(APIView):
#     """
#     Unified API endpoint to update pricing for multiple clusters or facilities.
#     Supports bulk operations with same response structure as individual APIs.
#     """
#     permission_classes = [AllowAny]  # Adjust permissions as needed
    
#     def put(self, request):
#         # Get request data
#         update_type = request.data.get('type')
#         updates = request.data.get('updates', [])
        
#         if not update_type or update_type not in ['cluster', 'facility']:
#             return Response(
#                 {"error": "type is required and must be 'cluster' or 'facility'"}, 
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         if not updates or not isinstance(updates, list):
#             return Response(
#                 {"error": "updates is required and must be a list"}, 
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         # Initialize aggregated response data
#         total_updated_records = 0
#         total_products_updated = 0
#         all_price_history_records = []
#         all_rejected_items = []
#         total_rejected_records_count = 0
        
#         # Get the current user (if authenticated)
#         current_user = request.user if hasattr(request, 'user') and request.user.is_authenticated else None
        
#         # Process each update
#         for update in updates:
#             try:
#                 if update_type == 'cluster':
#                     result = self._process_cluster_update(update, current_user)
#                 else:  # facility
#                     result = self._process_facility_update(update, current_user)
                
#                 # Aggregate results
#                 total_updated_records += result['total_updated_records']
#                 total_products_updated += result['total_products_updated']
#                 all_price_history_records.extend(result['price_history_records'])
#                 all_rejected_items.extend(result['rejected_items'])
#                 total_rejected_records_count += result['rejected_records_count']
                
#             except Exception as e:
#                 # If individual update fails, continue with others
#                 all_rejected_items.append({
#                     'error': str(e),
#                     'update': update
#                 })
#                 total_rejected_records_count += 1
#                 continue
        
#         if total_updated_records == 0:
#             return Response(
#                 {"error": "No records updated. Products may not have base prices set."}, 
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         # Create response message
#         entity_type = "clusters" if update_type == 'cluster' else "facilities"
#         message = f"Prices updated successfully for {len(updates)} {entity_type}"
        
#         return Response({
#             "success": True,
#             "message": message,
#             "type": update_type,
#             "total_updates_processed": len(updates),
#             "total_products_updated": total_products_updated,
#             "total_records_updated": total_updated_records,
#             "history_records_created": len(all_price_history_records),
#             "rejected_records_count": total_rejected_records_count,
#             "rejected_items": all_rejected_items,
#             "price_history_ids": [record.id for record in all_price_history_records]
#         }, status=status.HTTP_200_OK)
    
#     def _process_cluster_update(self, update, current_user):
#         """Process a single cluster update using existing logic"""
#         cluster_id = update.get('cluster_id')
#         margin = update.get('margin')
#         category_id = update.get('category_id')
#         brand_id = update.get('brand_id')
        
#         if not cluster_id:
#             raise Exception("cluster_id is required")
#         if not margin:
#             raise Exception("margin is required")
        
#         # Validate category_id if provided
#         if category_id:
#             try:
#                 from cms.models.category import Category
#                 Category.objects.get(id=category_id, is_active=True)
#             except Category.DoesNotExist:
#                 raise Exception(f"Category with ID {category_id} not found or inactive")
        
#         # Validate brand_id if provided
#         if brand_id:
#             try:
#                 from cms.models.category import Brand
#                 Brand.objects.get(id=brand_id, is_active=True)
#             except Brand.DoesNotExist:
#                 raise Exception(f"Brand with ID {brand_id} not found or inactive")
        
#         # Get the cluster
#         from cms.models.facility import Cluster
#         try:
#             cluster = Cluster.objects.get(id=cluster_id, is_active=True)
#         except Cluster.DoesNotExist:
#             raise Exception("Cluster not found or inactive")
        
#         # Get all facilities in the cluster
#         cluster_facilities = cluster.facilities.all()
#         if not cluster_facilities.exists():
#             raise Exception(f"No facilities found in cluster '{cluster.name}'")
        
#         # Determine which products to update based on filters
#         if category_id:
#             products_in_cluster = Product.objects.filter(
#                 category_id=category_id,
#                 is_active=True,
#                 is_published=True
#             )
#         elif brand_id:
#             products_in_cluster = Product.objects.filter(
#                 brand_id=brand_id,
#                 is_active=True,
#                 is_published=True
#             )
#         else:
#             cluster_inventory = FacilityInventory.objects.filter(
#                 facility__in=cluster_facilities,
#                 is_active=True
#             )
#             product_ids = cluster_inventory.values_list('product_variant__product_id', flat=True).distinct()
#             products_in_cluster = Product.objects.filter(
#                 id__in=product_ids,
#                 is_active=True,
#                 is_published=True
#             )
        
#         if not products_in_cluster.exists():
#             filter_name = "category" if category_id else "brand" if brand_id else "products"
#             raise Exception(f"No products found for the specified {filter_name} in cluster '{cluster.name}'")
        
#         # Update pricing for all products in the cluster
#         total_updated_records = 0
#         price_history_records = []
#         rejected_items = []
#         rejected_records_count = 0
#         total_products_updated = 0
        
#         for product in products_in_cluster:
#             product_variants = product.variants.filter(is_active=True)
#             product_updated_records = 0
            
#             for facility in cluster_facilities:
#                 for variant in product_variants:
#                     inventory, created = FacilityInventory.objects.get_or_create(
#                         facility=facility,
#                         product_variant=variant,
#                         defaults={'base_price': 0.0, 'selling_price': 0.0, 'is_active': True}
#                     )
                    
#                     current_selling_price = inventory.selling_price if inventory.selling_price and inventory.selling_price > 0 else (variant.base_price or 0.0)
#                     if current_selling_price and current_selling_price > 0:
#                         new_price = current_selling_price * (1 + margin / 100)
                        
#                         # Check if new price exceeds MRP
#                         variant_mrp = variant.mrp or 0.0
#                         if variant_mrp > 0 and new_price > variant_mrp:
#                             rejected_items.append({
#                                 'product_id': product.id,
#                                 'product_name': product.name,
#                                 'variant_id': variant.id,
#                                 'variant_name': variant.name,
#                                 'facility_id': facility.id,
#                                 'facility_name': facility.name,
#                                 'base_price': current_selling_price,
#                                 'calculated_price': new_price,
#                                 'mrp': variant_mrp,
#                                 'selling_price': current_selling_price,
#                                 'reason': 'calculated_price_exceeds_mrp'
#                             })
#                             rejected_records_count += 1
#                             continue
                        
#                         old_price = inventory.selling_price if inventory.selling_price > 0 else current_selling_price
                        
#                         # Update only selling price
#                         inventory.selling_price = new_price
#                         inventory.save()
                        
#                         # Create price history record
#                         price_history = ProductPriceHistory.objects.create(
#                             product=product,
#                             product_variant=variant,
#                             cluster=cluster,
#                             facility=facility,
#                             user=current_user,
#                             old_price=old_price,
#                             new_price=new_price,
#                             old_csp=old_price,
#                             new_csp=new_price,
#                             percentage_change=margin,
#                             change_type='bulk_cluster_update',
#                             change_reason=f'Bulk price update by {margin}% for cluster {cluster.name}'
#                         )
#                         price_history_records.append(price_history)
#                         product_updated_records += 1
#                         total_updated_records += 1
            
#             if product_updated_records > 0:
#                 total_products_updated += 1
        
#         return {
#             'total_updated_records': total_updated_records,
#             'total_products_updated': total_products_updated,
#             'price_history_records': price_history_records,
#             'rejected_items': rejected_items,
#             'rejected_records_count': rejected_records_count
#         }
    
#     def _process_facility_update(self, update, current_user):
#         """Process a single facility update using existing logic"""
#         facility_id = update.get('facility_id')
#         margin = update.get('margin')
#         category_id = update.get('category_id')
#         brand_id = update.get('brand_id')
        
#         if not facility_id:
#             raise Exception("facility_id is required")
#         if not margin:
#             raise Exception("margin is required")
        
#         # Validate category_id if provided
#         if category_id:
#             try:
#                 from cms.models.category import Category
#                 Category.objects.get(id=category_id, is_active=True)
#             except Category.DoesNotExist:
#                 raise Exception(f"Category with ID {category_id} not found or inactive")
        
#         # Validate brand_id if provided
#         if brand_id:
#             try:
#                 from cms.models.category import Brand
#                 Brand.objects.get(id=brand_id, is_active=True)
#             except Brand.DoesNotExist:
#                 raise Exception(f"Brand with ID {brand_id} not found or inactive")
        
#         # Get the facility
#         from cms.models.facility import Facility
#         try:
#             facility = Facility.objects.get(id=facility_id, is_active=True)
#         except Facility.DoesNotExist:
#             raise Exception("Facility not found or inactive")
        
#         # Determine which products to update based on filters
#         if category_id:
#             products_in_facility = Product.objects.filter(
#                 category_id=category_id,
#                 is_active=True,
#                 is_published=True
#             )
#         elif brand_id:
#             products_in_facility = Product.objects.filter(
#                 brand_id=brand_id,
#                 is_active=True,
#                 is_published=True
#             )
#         else:
#             facility_inventory = FacilityInventory.objects.filter(
#                 facility=facility,
#                 is_active=True
#             )
#             product_ids = facility_inventory.values_list('product_variant__product_id', flat=True).distinct()
#             products_in_facility = Product.objects.filter(
#                 id__in=product_ids,
#                 is_active=True,
#                 is_published=True
#             )
        
#         if not products_in_facility.exists():
#             filter_name = "category" if category_id else "brand" if brand_id else "products"
#             raise Exception(f"No products found for the specified {filter_name} in facility '{facility.name}'")
        
#         # Update pricing for all products in the facility
#         total_updated_records = 0
#         price_history_records = []
#         rejected_items = []
#         rejected_records_count = 0
#         total_products_updated = 0
        
#         for product in products_in_facility:
#             product_variants = product.variants.filter(is_active=True)
#             product_updated_records = 0
            
#             for variant in product_variants:
#                 inventory, created = FacilityInventory.objects.get_or_create(
#                     facility=facility,
#                     product_variant=variant,
#                     defaults={'base_price': 0.0, 'selling_price': 0.0, 'is_active': True}
#                 )
                
#                 current_selling_price = inventory.selling_price if inventory.selling_price and inventory.selling_price > 0 else (variant.base_price or 0.0)
#                 if current_selling_price and current_selling_price > 0:
#                     new_price = current_selling_price * (1 + margin / 100)
                    
#                     # Check if new price exceeds MRP
#                     variant_mrp = variant.mrp or 0.0
#                     if variant_mrp > 0 and new_price > variant_mrp:
#                         rejected_items.append({
#                             'product_id': product.id,
#                             'product_name': product.name,
#                             'variant_id': variant.id,
#                             'variant_name': variant.name,
#                             'facility_id': facility.id,
#                             'facility_name': facility.name,
#                             'base_price': current_selling_price,
#                             'calculated_price': new_price,
#                             'mrp': variant_mrp,
#                             'selling_price': current_selling_price,
#                             'reason': 'calculated_price_exceeds_mrp'
#                         })
#                         rejected_records_count += 1
#                         continue
                    
#                     old_price = inventory.selling_price if inventory.selling_price > 0 else current_selling_price
                    
#                     # Update only selling price
#                     inventory.selling_price = new_price
#                     inventory.save()
                    
#                     # Get cluster for price history (if facility is in a cluster)
#                     cluster = facility.clusters.first() if facility.clusters.exists() else None
                    
#                     # Create price history record
#                     price_history = ProductPriceHistory.objects.create(
#                         product=product,
#                         product_variant=variant,
#                         cluster=cluster,
#                         facility=facility,
#                         user=current_user,
#                         old_price=old_price,
#                         new_price=new_price,
#                         old_csp=old_price,
#                         new_csp=new_price,
#                         percentage_change=margin,
#                         change_type='bulk_facility_update',
#                         change_reason=f'Bulk price update by {margin}% for facility {facility.name}'
#                     )
#                     price_history_records.append(price_history)
#                     product_updated_records += 1
#                     total_updated_records += 1
            
#             if product_updated_records > 0:
#                 total_products_updated += 1
        
#         return {
#             'total_updated_records': total_updated_records,
#             'total_products_updated': total_products_updated,
#             'price_history_records': price_history_records,
#             'rejected_items': rejected_items,
#             'rejected_records_count': rejected_records_count
#         }


class ClusterPriceUpdateStatusView(APIView):
    """
    API endpoint to check if cluster prices have been updated recently.
    Shows who updated, when, and by what percentage.
    """
    permission_classes = [AllowAny]  # Adjust permissions as needed
    
    def post(self, request):
        # Validate request data
        cluster_id = request.data.get('cluster_id')
        if not cluster_id:
            return Response(
                {"error": "cluster_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the cluster
            from cms.models.facility import Cluster
            cluster = Cluster.objects.get(id=cluster_id, is_active=True)
        except Cluster.DoesNotExist:
            return Response(
                {"error": "Cluster not found or inactive"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the most recent price update for this cluster
        recent_update = ProductPriceHistory.objects.filter(
            cluster=cluster,
            change_type__in=['bulk_cluster_update', 'percentage_update']
        ).order_by('-creation_date').first()
        
        if not recent_update:
            return Response({
                "cluster_id": cluster_id,
                "cluster_name": cluster.name,
                "status": "not_updated",
                "message": f"No price updates found for cluster '{cluster.name}'",
                "last_update": None
            }, status=status.HTTP_200_OK)
        
        # Get all updates for this cluster (for detailed history)
        all_updates_queryset = ProductPriceHistory.objects.filter(
            cluster=cluster,
            change_type__in=['bulk_cluster_update', 'percentage_update']
        ).order_by('-creation_date')
        
        # Get last 10 updates for display
        recent_updates_list = list(all_updates_queryset[:10])
        
        # Get unique users who made updates
        update_users = {}
        for update in recent_updates_list:
            if update.user:
                user_id = update.user.id
                if user_id not in update_users:
                    update_users[user_id] = {
                        'id': update.user.id,
                        'username': update.user.username,
                        'name': f"{update.user.first_name} {update.user.last_name}".strip() or update.user.username
                    }
        
        # Calculate time since last update
        from django.utils import timezone
        now = timezone.now()
        time_diff = now - recent_update.creation_date
        
        # Format time difference
        if time_diff.days > 0:
            time_ago = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
        elif time_diff.seconds > 3600:
            hours = time_diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif time_diff.seconds > 60:
            minutes = time_diff.seconds // 60
            time_ago = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        else:
            time_ago = "Just now"
        
        # Get update statistics (use the full queryset, not sliced)
        total_updates = all_updates_queryset.count()
        unique_products = all_updates_queryset.values('product').distinct().count()
        unique_variants = all_updates_queryset.values('product_variant').distinct().count()
        
        # Get percentage changes used
        percentage_changes = list(all_updates_queryset.values_list('percentage_change', flat=True).distinct())
        
        return Response({
            "cluster_id": cluster_id,
            "cluster_name": cluster.name,
            "status": "updated",
            "message": f"Cluster '{cluster.name}' prices were last updated {time_ago}",
            "last_update": {
                "date": recent_update.creation_date.isoformat(),
                "time_ago": time_ago,
                "margin_update": recent_update.percentage_change,
                "updated_by": {
                    "id": recent_update.user.id if recent_update.user else None,
                    "username": recent_update.user.username if recent_update.user else "System",
                    "name": f"{recent_update.user.first_name} {recent_update.user.last_name}".strip() if recent_update.user and recent_update.user.first_name else (recent_update.user.username if recent_update.user else "System")
                },
                "change_type": recent_update.change_type,
                "change_reason": recent_update.change_reason
            },
            "update_statistics": {
                "total_updates": total_updates,
                "unique_products_updated": unique_products,
                "unique_variants_updated": unique_variants,
                "percentage_changes_used": percentage_changes,
                "update_users": list(update_users.values())
            },
            "recent_updates": [
                {
                    "id": update.id,
                    "date": update.creation_date.isoformat(),
                    "percentage_change": update.percentage_change,
                    "updated_by": {
                        "id": update.user.id if update.user else None,
                        "username": update.user.username if update.user else "System",
                        "name": f"{update.user.first_name} {update.user.last_name}".strip() if update.user and update.user.first_name else (update.user.username if update.user else "System")
                    },
                    "change_type": update.change_type,
                    "change_reason": update.change_reason,
                    "products_affected": update.product.name if update.product else "Unknown"
                }
                for update in recent_updates_list[:5]  # Last 5 updates
            ]
        }, status=status.HTTP_200_OK)


class SmartBrandBulkCreateProductsView(APIView):
    """
    POST /api/cms/products/smart-brand-bulk-create/
    Creates multiple products with smart brand assignment logic from file upload.
    Accepts CSV or Excel files with product data.
    Brand field accepts either ID (numeric) or name (text).
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        # Check if file is provided
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided. Please upload a CSV or Excel file."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        try:
            # Process file based on extension
            if file_extension == 'csv':
                products_data = self._process_csv_file(file)
            elif file_extension in ['xlsx', 'xls']:
                products_data = self._process_excel_file(file)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload CSV or Excel file."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not products_data:
                return Response(
                    {"error": "No valid data found in the file."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate all payloads as a list
            validator = SmartBrandProductSerializer(data=products_data, many=True)
            validator.is_valid(raise_exception=False)

            # collect index→errors
            errors = {
                idx: errs for idx, errs in enumerate(validator.errors) if errs
            }

            # pull out the valid payload dicts
            valid_items = [
                item for idx, item in enumerate(validator.validated_data)
                if idx not in errors
            ]

            created_products = []
            for item in valid_items:
                name = item['name']
                # if a product with this name exists, don't create – just return it
                existing = Product.objects.filter(name=name).first()
                if existing:
                    created_products.append(existing)
                else:
                    # truly new: use our single-item .create() to handle nested data
                    new_prod = SmartBrandProductSerializer().create(item)
                    # Track who created the product
                    new_prod.created_by = request.user
                    new_prod.updated_by = request.user
                    new_prod.save()
                    created_products.append(new_prod)

            # serialize output
            output_data = ProductDetailSerializer(created_products, many=True).data

            status_code = status.HTTP_201_CREATED if not errors else status.HTTP_207_MULTI_STATUS
            return Response({
                "message": f"Processed {len(products_data)} products from file '{file.name}'",
                "file_name": file.name,
                "total_processed": len(products_data),
                "successful_creates": len(created_products),
                "errors_count": len(errors),
                "created": output_data, 
                "errors": errors
            }, status=status_code)
            
        except Exception as e:
            return Response(
                {"error": f"Error processing file: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _process_csv_file(self, file):
        """Process CSV file and convert to product data format"""
        try:
            # Read CSV file
            csv_data = file.read().decode('utf-8')
            csv_reader = csv.DictReader(csv_data.splitlines())
            
            products_data = []
            for row in csv_reader:
                # Convert string values to appropriate types
                product_data = self._convert_row_to_product_data(row)
                products_data.append(product_data)
            
            return products_data
        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}")

    def _process_excel_file(self, file):
        """Process Excel file and convert to product data format"""
        try:
            # Reset file pointer to beginning
            file.seek(0)
            
            # Read Excel file with proper handling
            df = pd.read_excel(file, engine='openpyxl')
            
            # Debug: Print column names and first few rows
            print("Excel columns:", df.columns.tolist())
            print("First row data:", df.iloc[0].to_dict() if len(df) > 0 else "No data")
            
            # Convert NaN values to empty strings for string fields
            df = df.fillna('')
            
            products_data = []
            for _, row in df.iterrows():
                # Convert pandas row to dictionary
                row_dict = row.to_dict()
                product_data = self._convert_row_to_product_data(row_dict)
                products_data.append(product_data)
            
            return products_data
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")

    def _convert_row_to_product_data(self, row):
        """Convert a single row to product data format"""
        # Helper function to safely convert values
        def safe_int(value, default=0):
            try:
                if pd.isna(value) or value == '' or value is None:
                    return default
                return int(float(value))
            except (ValueError, TypeError):
                return default
        
        def safe_float(value, default=0.0):
            try:
                if pd.isna(value) or value == '' or value is None:
                    return default
                return float(value)
            except (ValueError, TypeError):
                return default
        
        def safe_str(value, default=''):
            try:
                if pd.isna(value) or value is None:
                    return default
                return str(value).strip()
            except (ValueError, TypeError):
                return default
        
        def safe_bool(value, default=True):
            try:
                if pd.isna(value) or value == '' or value is None:
                    return default
                return str(value).lower() in ['true', '1', 'yes', 'y']
            except (ValueError, TypeError):
                return default
        
        # These variables are now handled in the mapping section below
        
        # Map Excel columns to our expected fields
        # Product fields
        product_name = safe_str(row.get('Product Title', ''))
        product_description = safe_str(row.get('Product Description', ''))
        product_status = safe_str(row.get('Product Status', 'active'))
        product_brand = safe_str(row.get('Product Brand Id', ''))
        product_category = safe_int(row.get('Product Category Id', 1))
        product_tags = safe_str(row.get('Product Tags', ''))
        
        # Brand value is now working correctly
        
        # Variant fields - Updated to match your Excel column names
        variant_name = safe_str(row.get('Variant Title', ''))
        variant_sku = safe_str(row.get('Variant SKU', ''))
        variant_ean = safe_str(row.get('Variant EAN', ''))
        variant_net_qty = safe_str(row.get('Variant Net Qty', ''))  # Net Qty
        variant_base_price = safe_float(row.get('Variant Base Price', 0.0))  # Base Price
        variant_mrp = safe_float(row.get('Variant MRP', 0.0))  # MRP (Fixed: was 'Variant Mrp' before)
        variant_selling_price = safe_float(row.get('Variant Selling Price', 0.0))  # Selling Price
        # Variant options
        option1_name = safe_str(row.get('Variant Option 1 Name', ''))
        option1_value = safe_str(row.get('Variant Option 1 Value', ''))
        option2_name = safe_str(row.get('Variant Option 2 Name', ''))
        option2_value = safe_str(row.get('Variant Option 2 Value', ''))
        option3_name = safe_str(row.get('Variant Option 3 Name', ''))
        option3_value = safe_str(row.get('Variant Option 3 Value', ''))
        
        # Create combined variant name from all option values
        option_values = []
        weight_value = ""
        
        # Option values are now working correctly
        
        if option1_value and option1_value.strip():
            option_values.append(option1_value.strip())
        if option2_value and option2_value.strip():
            option_values.append(option2_value.strip())
        if option3_value and option3_value.strip():
            option_values.append(option3_value.strip())
            # Check if option 3 is weight
            if option3_name and option3_name.lower() == 'weight':
                weight_value = option3_value.strip()
                print(f"Weight detected: {weight_value}")
        
        # Use combined options as variant name, or fallback to original variant title
        combined_variant_name = " / ".join(option_values) if option_values else variant_name
        print(f"Combined variant name: {combined_variant_name}")
        print(f"Weight value: {weight_value}")
        
        # Image fields
        image1_url = safe_str(row.get('Product Image 1 Url', ''))
        image2_url = safe_str(row.get('Product Image 2 Url', ''))
        
        # Image URLs are working correctly
        
        # Convert status to boolean
        is_active = product_status.lower() in ['active', 'true', '1', 'yes']
        is_published = is_active  # Assuming active products are published
        
        # Structure the data for the bulk API
        product_data = {
            'name': product_name,
            'description': product_description,
            'tags': product_tags,
            'category': product_category,
            'brand': product_brand,
            'is_active': is_active,
            'is_published': is_published,
            'variants': [{
                'name': combined_variant_name,  # Use combined variant name
                'sku': variant_sku,
                'ean_number': variant_ean,
                'net_qty': variant_net_qty,
                'base_price': variant_base_price,
                'mrp': variant_mrp,
                'selling_price': variant_selling_price,
                'size': option1_value if option1_name.lower() == 'size' else option2_value if option2_name.lower() == 'size' else option3_value if option3_name.lower() == 'size' else '',
                'color': option1_value if option1_name.lower() == 'color' else option2_value if option2_name.lower() == 'color' else option3_value if option3_name.lower() == 'color' else '',
                'weight': weight_value,  # Add weight field
                'stock_quantity': 0  # Not provided in your Excel
            }] if combined_variant_name else [],
            'product_images': []
        }
        
        # Add images if they exist
        if image1_url:
            product_data['product_images'].append({
                'image': image1_url,
                'alt_text': f"{product_name} - Image 1",
                'priority': 1,
                'is_primary': True
            })
        
        if image2_url:
            product_data['product_images'].append({
                'image': image2_url,
                'alt_text': f"{product_name} - Image 2",
                'priority': 2,
                'is_primary': False
            })
        
        return product_data


# Size Chart Utility Functions
def handle_product_size_chart(product_variant, size_chart_data):
    """
    Handle size chart values for a product variant

    Args:
        product_variant: ProductVariant instance
        size_chart_data: Dictionary with size chart values

    Expected format:
    {
        "size_chart_values": [
            {
                "size": "M",  # AttributeValue.value or AttributeValue.id
                "measurements": {
                    "Chest": "36",
                    "Waist": "32",
                    "Shoulder": "18"
                }
            },
            {
                "size": "L",
                "measurements": {
                    "Chest": "38",
                    "Waist": "34",
                    "Shoulder": "19"
                }
            }
        ]
    }
    """
    if not size_chart_data or 'size_chart_values' not in size_chart_data:
        print(f"No size chart data provided for variant {product_variant.name}")
        return

    print(f"Processing size chart for variant: {product_variant.name}")

    # Get the size chart for the product's category
    try:
        size_chart = SizeChart.objects.get(
            category=product_variant.product.category,
            is_active=True
        )
    except SizeChart.DoesNotExist:
        print(f"No size chart configured for category '{product_variant.product.category.name}' (ID: {product_variant.product.category.id})")
        return  # No size chart configured for this category

    # Clear existing size chart values for this variant
    ProductSizeChartValue.objects.filter(product_variant=product_variant).delete()

    # Process each size and its measurements
    for size_data in size_chart_data['size_chart_values']:
        size_value = size_data.get('size')
        measurements = size_data.get('measurements', {})

        # Get the AttributeValue for this size
        try:
            if isinstance(size_value, int):
                size_attribute_value = AttributeValue.objects.get(
                    id=size_value,
                    attribute=size_chart.attribute
                )
            else:
                size_attribute_value = AttributeValue.objects.get(
                    value=size_value,
                    attribute=size_chart.attribute,
                    is_active=True
                )
        except AttributeValue.DoesNotExist:
            print(f"Size value '{size_value}' not found for attribute '{size_chart.attribute.name}'")
            continue  # Skip invalid size values

        # Create ProductSizeChartValue for each measurement
        for measurement_name, measurement_value in measurements.items():
            try:
                measurement = SizeMeasurement.objects.get(
                    size_chart=size_chart,
                    name=measurement_name,
                    is_active=True
                )

                ProductSizeChartValue.objects.create(
                    product_variant=product_variant,
                    size_attribute_value=size_attribute_value,
                    measurement=measurement,
                    value=str(measurement_value)
                )
            except SizeMeasurement.DoesNotExist:
                print(f"Measurement '{measurement_name}' not found for size chart '{size_chart.name}'")
                continue  # Skip invalid measurement names


def get_product_size_chart_values(product_variant):
    """
    Get size chart values for a product variant in a structured format

    Returns:
    {
        "size_chart_values": [
            {
                "size_id": 1,
                "size": "M",
                "measurements": {
                    "Chest": "36",
                    "Waist": "32"
                }
            }
        ]
    }
    """
    size_chart_values = ProductSizeChartValue.objects.filter(
        product_variant=product_variant
    ).select_related(
        'size_attribute_value',
        'measurement'
    ).order_by(
        'size_attribute_value__rank',
        'measurement__rank'
    )

    # Group by size
    size_data = {}
    for value in size_chart_values:
        size_key = value.size_attribute_value.id
        size_name = value.size_attribute_value.value

        if size_key not in size_data:
            size_data[size_key] = {
                'size_id': size_key,
                'size': size_name,
                'measurements': {}
            }

        size_data[size_key]['measurements'][value.measurement.name] = value.value

    return {
        'size_chart_values': list(size_data.values())
    }


class CategoryRequiredFieldsView(APIView):
    """
    API to get required fields for product creation based on category configuration
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        category_id = request.query_params.get('category_id')

        if not category_id:
            return Response({
                'error': 'category_id parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            return Response({
                'error': f'Category with ID {category_id} not found'
            }, status=status.HTTP_404_NOT_FOUND)

        required_fields = []
        optional_fields = []

        # 3. Check shelf life requirement
        if category.shelf_life_required:
            required_fields.append({
                'field_name': 'shelf_life',
                'required': True,
                'type': 'integer',
                'description': 'Shelf life days is required for this category'
            })

        # 4. Check size chart requirement
        try:
            from cms.models.setting import SizeChart
            size_chart = SizeChart.objects.get(category=category, is_active=True)
            
            # Get available measurements for this size chart
            measurements = size_chart.measurements.filter(is_active=True).order_by('rank', 'name')
            measurement_options = []
            for measurement in measurements:
                measurement_options.append({
                    'name': measurement.name,
                    'unit': measurement.unit,
                    'is_required': measurement.is_required,
                    'rank': measurement.rank
                })
            
            required_fields.append({
                'field_name': 'size_chart_values',
                'required': True,
                'type': 'size_chart',
                'description': f'Size chart values required for {size_chart.name}',
                'size_chart_id': size_chart.id,
                'size_chart_name': size_chart.name,
                'size_chart_attribute_id': size_chart.attribute.id,
                'size_chart_attribute_name': size_chart.attribute.name,
                'size_chart_attribute_type': size_chart.attribute.attribute_type,
                'measurements': measurement_options
            })
        except SizeChart.DoesNotExist:
            pass

        # 5. Check custom fields requirements from ProductType
        try:
            from cms.models.setting import ProductType
            product_type = ProductType.objects.get(category=category, is_active=True)

            for attribute in product_type.attributes.filter(is_required=True, is_active=True):
                attribute_data = {
                    'field_name': f'{attribute.name}',
                    'required': True,
                    'type': 'attribute',
                    'description': f'{attribute.name} is required',
                    'attribute_id': attribute.id,
                    'attribute_name': attribute.name,
                    'attribute_type': attribute.attribute_type
                }

                # Include option values for select-type attributes
                if attribute.attribute_type in ['select', 'multiselect']:
                    option_values = list(attribute.values.filter(is_active=True).values_list('value', flat=True))
                    if option_values:
                        attribute_data['options'] = option_values

                required_fields.append(attribute_data)

            # Also get optional attributes
            for attribute in product_type.attributes.filter(is_required=False, is_active=True):
                attribute_data = {
                    'field_name': f'{attribute.name}',
                    'required': False,
                    'type': 'attribute',
                    'description': f'{attribute.name} is optional',
                    'attribute_id': attribute.id,
                    'attribute_name': attribute.name,
                    'attribute_type': attribute.attribute_type
                }

                # Include option values for select-type attributes
                if attribute.attribute_type in ['select', 'multiselect']:
                    option_values = list(attribute.values.filter(is_active=True).values_list('value', flat=True))
                    if option_values:
                        attribute_data['options'] = option_values

                optional_fields.append(attribute_data)

        except ProductType.DoesNotExist:
            pass

        # 6. Check custom fields from tabs/sections based on category
        try:
            from cms.models.setting import CustomTab
            tabs = CustomTab.objects.filter(category=category, is_active=True)

            for tab in tabs:
                for section in tab.sections.filter(is_active=True):
                    # Required custom fields
                    for field in section.fields.filter(is_required=True, is_active=True):
                        field_data = {
                            'field_name': f'{field.name}',
                            'required': True,
                            'type': 'custom_field',
                            'description': f'{field.label} is required',
                            'custom_field_id': field.id,
                            'field_type': field.field_type,
                            'field_label': field.label,
                            'field_name_key': field.name,
                        }

                        # Include options for fields that support them
                        if field.field_type in ['select', 'multiselect', 'radio', 'checkbox'] and field.options:
                            field_data['options'] = field.options

                        required_fields.append(field_data)

                    # Optional custom fields
                    for field in section.fields.filter(is_required=False, is_active=True):
                        field_data = {
                            'field_name': f'{field.name}',
                            'required': False,
                            'type': 'custom_field',
                            'description': f'{field.label} is optional',
                            'custom_field_id': field.id,
                            'field_type': field.field_type,
                            'field_label': field.label,
                            'field_name_key': field.name,
                        }

                        # Include options for fields that support them
                        if field.field_type in ['select', 'multiselect', 'radio', 'checkbox'] and field.options:
                            field_data['options'] = field.options

                        optional_fields.append(field_data)
        except Exception:
            pass

        response_data = {
            'category_id': category.id,
            'category_name': category.name,
            'required_fields': required_fields,
            'optional_fields': optional_fields,
            'total_required_fields': len(required_fields),
            'total_optional_fields': len(optional_fields)
        }

        return Response(response_data, status=status.HTTP_200_OK)


class GS1APIView(APIView):
    """
    API endpoint to call GS1 API and return the same response
    Usage: GET /api/gs1/?ean=1234567890123
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Get EAN/GTIN information via query parameter
        Usage: GET /api/gs1/?ean=1234567890123
        """
        ean_number = request.query_params.get('ean')

        if not ean_number:
            return Response({
                'error': 'ean parameter is required',
                'message': 'Please provide an EAN/GTIN number as query parameter'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check if EAN already exists in database
        existing_variant = ProductVariant.objects.filter(ean_number=ean_number).first()
        if existing_variant:
            return Response({
                'status': False,
                'error': 'EAN already exists',
                'message': f'EAN {ean_number} already exists in the database',
                'existing_product': {
                    'variant_id': existing_variant.id,
                    'variant_name': existing_variant.name,
                    'variant_sku': existing_variant.sku,
                    'product_id': existing_variant.product.id,
                    'product_name': existing_variant.product.name,
                }
            }, status=status.HTTP_409_CONFLICT)

        # Get GS1 API configuration
        import os
        import json
        import requests

        GS1_API_URL = os.environ.get("GS1_API_URL", "https://api.gs1datakart.org/console/retailer/products")
        GS1_API_TOKEN = os.environ.get("GS1_API_TOKEN", "")

        if not GS1_API_TOKEN:
            return Response({
                'error': 'GS1 API token not configured',
                'message': 'Please configure GS1_API_TOKEN in environment variables'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            # Prepare request parameters (GS1 API expects array format)
            gtin_param = json.dumps([ean_number])

            # Make request to GS1 API
            response = requests.get(
                GS1_API_URL,
                params={'gtin': gtin_param, 'status': 'published'},
                headers={'Authorization': f'Bearer {GS1_API_TOKEN}'}
            )

            # Return the same response from GS1 API
            try:
                response_data = response.json()
                # Add database check result to response
                response_data['db_check'] = {
                    'exists': False,
                    'message': 'EAN does not exist in database'
                }
                return Response(response_data, status=response.status_code)
            except ValueError:
                # If response is not JSON, return as text
                return Response({
                    'data': response.text,
                    'status_code': response.status_code,
                    'db_check': {
                        'exists': False,
                        'message': 'EAN does not exist in database'
                    }
                }, status=response.status_code)

        except requests.exceptions.RequestException as e:
            return Response({
                'error': 'Failed to call GS1 API',
                'message': str(e)
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        except Exception as e:
            return Response({
                'error': 'Internal server error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CollectionExportView(APIView):
    """
    GET /api/collections/export/
    Exports collections data in Excel format.

    Query Parameters:
    - search: Search in collection name
    - status: Filter by active/inactive status
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum records (default: 1000, max: 10000)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import datetime
        from cms.utils.filter import CollectionFilter

        # Apply filters using the same filter class as CollectionViewSet
        from cms.models.product import Collection
        filter_instance = CollectionFilter(request.GET, queryset=Collection.objects.prefetch_related('products'))
        queryset = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | Q(description__icontains=search_query)
            )

        # Apply ordering same as CollectionViewSet
        queryset = queryset.order_by('name')

        # Limit results for performance
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)
        queryset = queryset[:export_limit]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Collections"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'ID', 'Name', 'Description', 'Status', 'Has Image',
            'Total Products', 'Created Date', 'Updated Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, collection in enumerate(queryset, 2):
            product_count = collection.products.count()

            data = [
                collection.id,
                collection.name,
                collection.description or '',
                'Active' if collection.is_active else 'Inactive',
                'Yes' if collection.image else 'No',
                product_count,
                collection.creation_date.strftime('%Y-%m-%d %H:%M:%S') if collection.creation_date else '',
                collection.updation_date.strftime('%Y-%m-%d %H:%M:%S') if collection.updation_date else ''
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [1, 6]:  # ID and Product count columns
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="collections_export_{timestamp}.xlsx"'

        wb.save(response)
        return response




class OverridePriceView(APIView):
    """
    Override pricing API with cluster/facility targeting and MRP validation.
    Supports discovery mode (no margin) and execution mode (with margin).
    """
    permission_classes = [AllowAny]
    
    def post(self, request):
        cluster_ids = request.data.get('cluster_ids', [])
        facility_ids = request.data.get('facility_ids', [])
        variant_ids = request.data.get('variant_ids', [])
        category_ids = request.data.get('category_ids', [])
        brand_ids = request.data.get('brand_ids', [])
        margin = request.data.get('margin')
        type_param = request.data.get('type')
        skip_price_history = request.data.get('skip_price_history', False)  # Enable history by default
        max_variants = request.data.get('max_variants', 5000)  # Reasonable default for fast response
        product_name = request.data.get('product_name', '')  # Filter by product name
        variant_name = request.data.get('variant_name', '')  # Filter by variant name
        
        # Pagination parameters - support both query params and body params
        page = request.data.get('page') or request.query_params.get('page', 1)
        page_size = request.data.get('page_size') or request.query_params.get('page_size', 50)
        
        # Validate pagination parameters
        try:
            page = int(page)
            page_size = int(page_size)
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 1000:  # Limit max page size
                page_size = 50
        except (ValueError, TypeError):
            page = 1
            page_size = 50
        
        # Validate required fields - either cluster_ids or facility_ids is required
        if not cluster_ids and not facility_ids:
            return Response(
                {"error": "Either cluster_ids or facility_ids is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate clusters exist (only if cluster_ids provided)
        clusters = Cluster.objects.none()
        if cluster_ids:
            try:
                clusters = Cluster.objects.filter(id__in=cluster_ids, is_active=True)
                if clusters.count() != len(cluster_ids):
                    return Response(
                        {"error": "One or more cluster IDs not found or inactive"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except Exception as e:
                return Response(
                    {"error": f"Invalid cluster_ids: {str(e)}"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Discovery mode - return available options
        if margin is None:
            return self._discovery_mode(clusters, facility_ids, variant_ids, category_ids, brand_ids, product_name, variant_name, page, page_size)
        
        # Validate execution mode parameters
        if variant_ids and type_param:
            return Response(
                {"error": "Cannot provide both variant_ids and type parameter. Use either variant_ids for specific variants or type: 'all' for all variants."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not variant_ids and not type_param:
            return Response(
                {"error": "Either variant_ids or type: 'all' is required when margin is provided."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if type_param and type_param != 'all':
            return Response(
                {"error": "Type must be 'all' when provided."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Execution mode - update prices
        try:
            result = self._execution_mode(clusters, facility_ids, variant_ids, category_ids, brand_ids, margin, type_param, product_name, variant_name, page, page_size, request.user, skip_price_history, max_variants)
            return result
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {"error": f"Internal server error: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _discovery_mode(self, clusters, facility_ids, variant_ids, category_ids, brand_ids, product_name, variant_name, page, page_size):
        """Return available facilities, variants, categories, and brands for the clusters"""
        # Prioritize facility_ids over cluster filtering
        if facility_ids:
            cluster_facilities = Facility.objects.filter(
                id__in=facility_ids, is_active=True
            )
        else:
            cluster_facilities = Facility.objects.filter(
                clusters__in=clusters, is_active=True
            ).distinct()
        
        if not cluster_facilities.exists():
            return Response(
                {"error": "No active facilities found in the specified clusters"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get variants in these facilities
        facility_inventories = FacilityInventory.objects.filter(
            facility__in=cluster_facilities, is_active=True
        ).select_related('product_variant__product')
        
        inventory_variant_ids = facility_inventories.values_list('product_variant_id', flat=True).distinct()
        variants = ProductVariant.objects.filter(id__in=inventory_variant_ids, is_active=True, is_published=True)
        
        # Filter by specific variants if provided
        if variant_ids:
            variants = variants.filter(id__in=variant_ids)

        # Apply filtering based on product name or variant name
        if product_name:
            variants = variants.filter(product__name__icontains=product_name)

        if variant_name:
            variants = variants.filter(name__icontains=variant_name)
        
        if not variants.exists():
            return Response(
                {"error": "No variants found matching the specified criteria"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Apply other filters through products
        if category_ids:
            variants = variants.filter(product__category_id__in=category_ids)
        if brand_ids:
            variants = variants.filter(product__brand_id__in=brand_ids)
        
        total_variants = variants.count()
        
        # Apply pagination
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_variants = variants[start_index:end_index]
        
        # Fetch selling prices for paginated variants per facility
        discovery_inventories = FacilityInventory.objects.filter(
            facility__in=cluster_facilities,
            product_variant__in=paginated_variants,
            is_active=True
        ).select_related('facility', 'product_variant')
        
        # Build a lookup: variant_id -> list of facility price dicts
        variant_id_to_facility_prices = {}
        for inv in discovery_inventories:
            facility_prices = variant_id_to_facility_prices.setdefault(inv.product_variant_id, [])
            facility_prices.append({
                'facility_id': inv.facility_id,
                'facility_name': getattr(inv.facility, 'name', ''),
                'selling_price': inv.selling_price or 0
            })

        # Compose variants payload with base_price, mrp, and selling prices per facility
        variants_payload = []
        
        # Fetch latest price history (margin/user) per variant within selected facilities
        last_history_by_variant = {}
        histories = ProductPriceHistory.objects.filter(
            product_variant_id__in=[v.id for v in paginated_variants],
            facility__in=cluster_facilities
        ).select_related('user').order_by('product_variant_id', '-creation_date')
        seen_variant_ids = set()
        for h in histories:
            if h.product_variant_id in seen_variant_ids:
                continue
            seen_variant_ids.add(h.product_variant_id)
            last_history_by_variant[h.product_variant_id] = {
                'margin': h.percentage_change,
                'user': ({
                    'id': getattr(h.user, 'id', None),
                    'name': getattr(h.user, 'name', None) or getattr(h.user, 'username', None),
                    'email': getattr(h.user, 'email', None)
                } if getattr(h, 'user', None) else None),
                'timestamp': h.creation_date
            }
        for v in paginated_variants:
            variants_payload.append({
                'id': v.id,
                'name': v.name,
                'sku': v.sku,
                'product_name': v.product.name,
                'product_id': v.product.id,
                'base_price': v.base_price or 0,
                'mrp': v.mrp or 0,
                'selling_prices': variant_id_to_facility_prices.get(v.id, []),
                'last_update': last_history_by_variant.get(v.id)
            })

        # Get categories and brands
        product_ids = variants.values_list('product_id', flat=True).distinct()
        products = Product.objects.filter(id__in=product_ids, is_active=True, is_published=True)
        
        categories = Category.objects.filter(
            products__in=products, is_active=True
        ).distinct().values('id', 'name')
        
        brands = Brand.objects.filter(
            products__in=products, is_active=True
        ).distinct().values('id', 'name')
        
        # Calculate pagination info
        total_pages = (total_variants + page_size - 1) // page_size
        has_next = page < total_pages
        has_previous = page > 1
        
        return Response({
            "mode": "discovery",
            "clusters": [{"id": c.id, "name": c.name} for c in clusters],
            "facilities": [{"id": f.id, "name": f.name, "city": f.city} for f in cluster_facilities],
            "variants": variants_payload,
            "categories": list(categories),
            "brands": list(brands),
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_variants": total_variants,
                "total_pages": total_pages,
                "has_next": has_next,
                "has_previous": has_previous,
                "next_page": page + 1 if has_next else None,
                "previous_page": page - 1 if has_previous else None
            },
            "total_variants": total_variants,
            "total_variants_count": total_variants,
            "variants_count": total_variants,
            "message": "Provide margin percentage and either variant_ids or type: 'all' to execute price updates"
        })
    
    def _execution_mode(self, clusters, facility_ids, variant_ids, category_ids, brand_ids, margin, type_param, product_name, variant_name, page, page_size, user, skip_price_history=False, max_variants=10000):
        """Execute price updates with MRP validation"""
        # Prioritize facility_ids over cluster filtering
        if facility_ids:
            # If facility_ids provided, use them directly (ignore cluster filtering)
            cluster_facilities = Facility.objects.filter(
                id__in=facility_ids, is_active=True
            )
        else:
            # Only cluster_ids provided, get facilities from clusters
            cluster_facilities = Facility.objects.filter(
                clusters__in=clusters, is_active=True
            ).distinct()
        
        if not cluster_facilities.exists():
            return Response(
                {"error": "No active facilities found"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get variants to update with optimized queries
        facility_inventories = FacilityInventory.objects.filter(
            facility__in=cluster_facilities, is_active=True
        ).select_related('product_variant__product', 'product_variant__product__category', 'product_variant__product__brand')
        
        inventory_variant_ids = facility_inventories.values_list('product_variant_id', flat=True).distinct()
        variants = ProductVariant.objects.filter(
            id__in=inventory_variant_ids, 
            is_active=True, 
            is_published=True
        ).select_related('product', 'product__category', 'product__brand')
        
        # Apply variant filtering
        if variant_ids:
            variants = variants.filter(id__in=variant_ids)
        elif type_param == 'all':
            # Get all variants matching other filters
            pass
        if product_name:
            variants = variants.filter(product__name__icontains=product_name)
        if variant_name:
            variants = variants.filter(name__icontains=variant_name)
        
        # Apply other filters through products
        if category_ids:
            variants = variants.filter(product__category_id__in=category_ids)
        if brand_ids:
            variants = variants.filter(product__brand_id__in=brand_ids)
        
        if not variants.exists():
            return Response(
                {"error": "No variants found matching the specified criteria"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get total count before pagination
        total_variants = variants.count()
        
        # Apply pagination only for discovery mode or when not using type: "all"
        if type_param == 'all':
            # For type: "all", process ALL variants (no limit)
            variants_to_process = variants
            print(f"Processing ALL {variants.count()} variants (type: 'all')")
        else:
            # Apply pagination for specific variant_ids
            start_index = (page - 1) * page_size
            end_index = start_index + page_size
            variants_to_process = variants[start_index:end_index]
        
        # Update prices with bulk operations for maximum performance
        updated_variants = []
        rejected_variants = []
        total_updated = 0
        total_rejected = 0
        total_processed = 0  # Track total variants that got a decision
        
        # Get all inventories at once to avoid N+1 queries
        inventories = FacilityInventory.objects.filter(
            facility__in=cluster_facilities,
            product_variant__in=variants_to_process,
            is_active=True
        ).select_related('facility', 'product_variant', 'product_variant__product')
        
        # Create lookup dictionary for faster access
        inventory_lookup = {}
        for inv in inventories:
            key = f"{inv.facility_id}_{inv.product_variant_id}"
            inventory_lookup[key] = inv
        
        # Also create a set of variant IDs that have inventory records for quick lookup
        variants_with_inventory = set()
        for inv in inventories:
            variants_with_inventory.add(inv.product_variant_id)
        
        # Prepare bulk updates
        inventories_to_update = []
        price_history_records = []
        
        # Process variants in smaller batches for better performance and memory management
        batch_size = 100  # Smaller batch size for faster processing
        variant_list = list(variants_to_process)
        total_variants_to_process = len(variant_list)
        
        print(f"Processing {total_variants_to_process} variants in batches of {batch_size}")
        
        for i in range(0, len(variant_list), batch_size):
            batch_variants = variant_list[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (total_variants_to_process + batch_size - 1) // batch_size
            
            print(f"Processing batch {batch_num}/{total_batches} ({len(batch_variants)} variants)")
            
            for variant in batch_variants:
                variant_updated = 0
                variant_rejected = 0
                variant_selling_prices = []  # Track selling prices for this variant
                
                # Check if this variant has any inventory records at all
                if variant.id not in variants_with_inventory:
                    # No inventory records for this variant in any facility - reject
                    if len(rejected_variants) < 100:
                        rejected_variants.append({
                            'variant_id': variant.id,
                            'variant_name': variant.name,
                            'sku': variant.sku,
                            'product_id': variant.product.id,
                            'product_name': variant.product.name,
                            'facility_id': cluster_facilities.first().id if cluster_facilities.exists() else None,
                            'facility_name': cluster_facilities.first().name if cluster_facilities.exists() else 'N/A',
                            'base_price': variant.base_price or 0,
                            'selling_price': 0,
                            'mrp': variant.mrp or 0,
                            'calculated_price': 0,
                            'reason': 'no_inventory_records'
                        })
                    variant_rejected += 1
                    total_rejected += 1
                    total_processed += 1
                    continue
                
                for facility in cluster_facilities:
                    key = f"{facility.id}_{variant.id}"
                    inventory = inventory_lookup.get(key)
                    
                    if not inventory:
                        # No inventory record for this specific facility - reject
                        if len(rejected_variants) < 100:
                            rejected_variants.append({
                                'variant_id': variant.id,
                                'variant_name': variant.name,
                                'sku': variant.sku,
                                'product_id': variant.product.id,
                                'product_name': variant.product.name,
                                'facility_id': facility.id,
                                'facility_name': facility.name,
                                'base_price': variant.base_price or 0,
                                'selling_price': 0,
                                'mrp': variant.mrp or 0,
                                'calculated_price': 0,
                                'reason': 'no_inventory_record'
                            })
                        variant_rejected += 1
                        continue
                    
                    # Get current prices
                    current_selling_price = inventory.selling_price or 0
                    base_price = variant.base_price or 0
                    variant_mrp = variant.mrp or 0
                    
                    # Determine which price to use for calculation
                    if current_selling_price > 0:
                        current_price = current_selling_price
                    elif base_price > 0:
                        current_price = base_price
                    else:
                        # No valid price - reject this variant
                        if len(rejected_variants) < 100:
                            rejected_variants.append({
                                'variant_id': variant.id,
                                'variant_name': variant.name,
                                'sku': variant.sku,
                                'product_id': variant.product.id,
                                'product_name': variant.product.name,
                                'facility_id': facility.id,
                                'facility_name': facility.name,
                                'base_price': base_price,
                                'selling_price': current_selling_price,
                                'mrp': variant_mrp,
                                'calculated_price': 0,
                                'reason': 'no_valid_price'
                            })
                        variant_rejected += 1
                        continue
                    
                    new_price = base_price * (1 + margin / 100)
                    
                    # MRP validation
                    if variant_mrp > 0 and new_price > variant_mrp:
                        # Only store first 100 rejected items to limit response size
                        if len(rejected_variants) < 100:
                            rejected_variants.append({
                                'variant_id': variant.id,
                                'variant_name': variant.name,
                                'sku': variant.sku,
                                'product_id': variant.product.id,
                                'product_name': variant.product.name,
                                'facility_id': facility.id,
                                'facility_name': facility.name,
                                'base_price': base_price,
                                'selling_price': current_selling_price,
                                'mrp': variant_mrp,
                                'calculated_price': new_price,
                                'reason': 'calculated_price_exceeds_mrp'
                            })
                        variant_rejected += 1
                        continue
                    
                    # Prepare for bulk update
                    old_price = inventory.selling_price
                    inventory.selling_price = new_price
                    inventories_to_update.append(inventory)
                    
                    # Track selling price for this variant
                    variant_selling_prices.append({
                        'facility_id': facility.id,
                        'facility_name': facility.name,
                        'old_selling_price': old_price or 0,
                        'new_selling_price': new_price
                    })
                    
                    # Prepare price history record for bulk creation (only if not skipped)
                    if not skip_price_history:
                        price_history_records.append(ProductPriceHistory(
                            product=variant.product,
                            product_variant=variant,
                            cluster=clusters.first() if clusters.exists() else None,
                            facility=facility,
                            user=user if user and user.is_authenticated else None,
                            old_price=old_price or current_price,
                            new_price=new_price,
                            old_csp=old_price or current_price,
                            new_csp=new_price,
                            percentage_change=margin,
                            change_type='override_price_update',
                            change_reason=f'Override price update by {margin}%'
                        ))
                    
                    variant_updated += 1
                    total_updated += 1
                
                # Always add variant to results if it was processed (updated or rejected)
                if variant_updated > 0 or variant_rejected > 0:
                    # Only store first 100 items to limit response size
                    if len(updated_variants) < 100:
                        updated_variants.append({
                            'variant_id': variant.id,
                            'variant_name': variant.name,
                            'sku': variant.sku,
                            'product_id': variant.product.id,
                            'product_name': variant.product.name,
                            'base_price': variant.base_price or 0,
                            'mrp': variant.mrp or 0,
                            'selling_prices': variant_selling_prices if variant_selling_prices else [],
                            'facilities_updated': variant_updated,
                            'facilities_rejected': variant_rejected,
                            'status': 'updated' if variant_updated > 0 else 'rejected'
                        })
                    total_rejected += variant_rejected
                    total_processed += 1
            
            # Perform bulk updates after each batch to prevent memory issues
            if inventories_to_update:
                print(f"Updating {len(inventories_to_update)} inventories in batch {batch_num}")
                FacilityInventory.objects.bulk_update(
                    inventories_to_update, 
                    ['selling_price'], 
                    batch_size=100
                )
                
                # Bulk create price history records (only if not skipped)
                if not skip_price_history and price_history_records:
                    print(f"Creating {len(price_history_records)} price history records in batch {batch_num}")
                    ProductPriceHistory.objects.bulk_create(
                        price_history_records, 
                        batch_size=100
                    )
                
                # Clear lists for next batch
                inventories_to_update.clear()
                price_history_records.clear()
        
        print(f"Completed processing all {total_variants_to_process} variants")
        
        # Calculate pagination info
        if type_param == 'all':
            # For type: "all", show that all variants were processed
            total_pages = 1
            has_next = False
            has_previous = False
        else:
            # Normal pagination for specific variant_ids
            total_pages = (total_variants + page_size - 1) // page_size
            has_next = page < total_pages
            has_previous = page > 1
        
        return Response({
            "success": True,
            "mode": "execution",
            "margin_applied": margin,
            "clusters_processed": [{"id": c.id, "name": c.name} for c in clusters],
            "facilities_processed": [{"id": f.id, "name": f.name} for f in cluster_facilities],
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_variants": total_variants,
                "total_pages": total_pages,
                "has_next": has_next,
                "has_previous": has_previous,
                "next_page": page + 1 if has_next else None,
                "previous_page": page - 1 if has_previous else None
            },
            "summary": {
                "total_variants_found": total_variants,
                "total_variants_processed": total_processed,
                "total_variants_updated": total_updated,
                "total_variants_rejected": total_rejected,
                "total_variants_skipped": len(variants_to_process) - total_processed,
                "variants_count": total_variants,
                "price_history_records_created": total_updated if not skip_price_history else 0
            },
            "updated_variants": updated_variants,
            "rejected_variants": rejected_variants,
            "price_history_enabled": not skip_price_history,
            "message": f"Price override completed. {total_updated} variants updated, {total_rejected} variants rejected. Price history: {'Enabled' if not skip_price_history else 'Disabled'}. (Note: Results limited to first 100 items for performance)"
        })


class ComboProductViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing combo products with comprehensive filtering.

    List: GET /api/combo-products/
    Retrieve: GET /api/combo-products/{id}/
    Create: POST /api/combo-products/
    Update: PUT /api/combo-products/{id}/
    Partial Update: PATCH /api/combo-products/{id}/
    Delete: DELETE /api/combo-products/{id}/

    Filters:
    - is_active: Filter by active status (true/false)
    - name: Filter by combo name (contains)
    - description: Filter by description (contains)
    - combo_variant_sku: Filter by combo variant SKU (contains)
    - combo_variant_name: Filter by combo variant name (contains)
    - product: Filter by product ID
    - product_name: Filter by product name (contains)
    - category: Filter by category ID
    - brand: Filter by brand ID
    - variant_is_active: Filter by combo variant active status
    - variant_is_published: Filter by combo variant published status
    - contains_variant: Filter combos containing a specific variant ID
    - created_after: Filter by creation date (>=)
    - created_before: Filter by creation date (<=)
    - updated_after: Filter by update date (>=)
    - updated_before: Filter by update date (<=)

    Search: name, combo_variant__name, combo_variant__sku, description
    Ordering: creation_date, updation_date, name, is_active
    """
    queryset = ComboProduct.objects.all()
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = ComboProductFilter
    search_fields = ['name', 'combo_variant__name', 'combo_variant__sku', 'description']
    ordering_fields = ['creation_date', 'updation_date', 'name', 'is_active']
    ordering = ['-creation_date']

    def get_queryset(self):
        """Optimize queries with select_related and prefetch_related"""
        queryset = super().get_queryset()
        return queryset.select_related('combo_variant', 'combo_variant__product').prefetch_related(
            'combo_items__product_variant__product',
            'combo_items__product_variant__images'
        )

    def get_serializer_class(self):
        """Use different serializers for list and create/update actions"""
        if self.action == 'list' or self.action == 'retrieve':
            return ComboProductListSerializer
        return ComboProductCreateSerializer

    def perform_create(self, serializer):
        """Set the combo variant's is_combo flag when creating"""
        combo_product = serializer.save()
        # Ensure the combo_variant has is_combo set to True
        if combo_product.combo_variant and not combo_product.combo_variant.is_combo:
            combo_product.combo_variant.is_combo = True
            combo_product.combo_variant.save()

    def destroy(self, request, *args, **kwargs):
        """Custom delete to unset is_combo flag on the variant"""
        instance = self.get_object()
        combo_variant = instance.combo_variant

        # Delete the combo product
        self.perform_destroy(instance)

        # Unset is_combo flag on the variant
        if combo_variant:
            combo_variant.is_combo = False
            combo_variant.save()

        return Response(status=status.HTTP_204_NO_CONTENT)

