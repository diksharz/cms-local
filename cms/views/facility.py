from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions, AllowAny
from cms.models.facility import Facility, Cluster, FacilityInventory
from cms.models.product import ProductVariant, Product
from cms.serializers.facility import (
    FacilitySerializer, ClusterSerializer, ClusterListSerializer, FacilityInventorySerializer,
    ProductListSerializer, FacilityInventoryItemSerializer
)
from rest_framework.filters import SearchFilter, OrderingFilter
from cms.utils.pagination import CustomPageNumberPagination
from rest_framework.response import Response
from rest_framework.decorators import action
from cms.utils.filter import (
    FacilityFilter,
    ProductFilter,
    ClusterFilter
)
from django_filters import rest_framework as filters
from django.db.models import Prefetch
from rest_framework import status
from django.http import HttpResponse
from django.db.models import Q
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView



class FacilityViewSet(viewsets.ModelViewSet):
    queryset = Facility.objects.all()
    serializer_class = FacilitySerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends  = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class  = FacilityFilter
    search_fields = ['name', 'facility_type']
    ordering_fields = ['name', 'facility_type', 'city', 'is_active', 'created_at']
    ordering = ['name'] 

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculate counts
        active_count = queryset.filter(is_active=True).count()
        inactive_count = queryset.filter(is_active=False).count()
        
        # Paginate
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['total_active_count'] = active_count
            response.data['total_inactive_count'] = inactive_count
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'count': queryset.count(),
            'total_active_count': active_count,
            'total_inactive_count': inactive_count
        })

    def perform_create(self, serializer):
        serializer.save()


class ClusterViewSet(viewsets.ModelViewSet):
    queryset = Cluster.objects.all()
    # serializer_class = ClusterSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = ClusterFilter
    search_fields = ['name', 'region']
    ordering_fields = ['name', 'region', 'is_active', 'creation_date', 'updation_date']
    ordering = ['name']

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return ClusterListSerializer
        else:
            return ClusterSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculate counts
        active_count = queryset.filter(is_active=True).count()
        inactive_count = queryset.filter(is_active=False).count()
        
        # Paginate
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['total_active_count'] = active_count
            response.data['total_inactive_count'] = inactive_count
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'count': queryset.count(),
            'total_active_count': active_count,
            'total_inactive_count': inactive_count
        })
            

class FacilityInventoryViewSet(viewsets.ModelViewSet):
    queryset = FacilityInventory.objects.all()
    serializer_class = FacilityInventorySerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends = (SearchFilter,)
    search_fields = ['facility', 'product_variant']

    def get_queryset(self):
        """
        Optionally restricts the returned facility inventories to a given facility.
        """
        queryset = super().get_queryset()
        
        # Check if 'facility_id' is passed as a query parameter
        facility_id = self.request.query_params.get('facility_id', None)
        
        if facility_id:
            queryset = queryset.filter(facility_id=facility_id)  # Filter by facility_id

        return queryset

    # def perform_create(self, serializer):
    #     serializer.save()

    @action(detail=False, methods=['post'], url_path='add')
    def bulk_create(self, request):
        """
        Handles the bulk creation of FacilityInventory entries for a given facility.
        Expects a list of product_variant IDs and creates them in bulk.
        Avoids creating duplicates if the combination of facility_id and product_variant_id exists.
        """
        facility_id = request.data.get('facility')
        product_variant_ids = request.data.get('product_variant', [])

        if not facility_id:
            return Response({'error': 'Facility ID is required'}, status=400)

        if not product_variant_ids:
            return Response({'error': 'Product variant IDs are required'}, status=400)

        # Fetch the facility object (optional: validate existence)
        try:
            facility = Facility.objects.get(id=facility_id)
        except Facility.DoesNotExist:
            return Response({'error': 'Facility not found'}, status=400)

        # List to hold FacilityInventory objects to be created
        facility_inventories = []
        existing_combinations = FacilityInventory.objects.filter(
            facility_id=facility_id, 
            product_variant_id__in=product_variant_ids
        ).values_list('product_variant_id', flat=True)

        for product_variant_id in product_variant_ids:
            if product_variant_id in existing_combinations:
                # Skip if this combination already exists
                continue

            try:
                product_variant = ProductVariant.objects.get(id=product_variant_id)
            except ProductVariant.DoesNotExist:
                return Response({'error': f'Product variant with ID {product_variant_id} not found'}, status=400)

            # Create FacilityInventory instance
            facility_inventory = FacilityInventory(
                facility=facility,
                product_variant=product_variant,
                stock=0,  # Default to 0, can be updated later
                tax=None,  # Default tax, can be added if needed
                base_price=0.0,  # Default base price, can be updated later
                mrp=0.0,  # Default MRP, can be updated
                selling_price=0.0,  # Default selling price, can be updated
                cust_discount=None,  # Optional customer discount
                max_purchase_limit=None,  # Optional
                outofstock_threshold=None,  # Optional
                status='Active',  # Default status
                is_active=True  # Default to active
            )

            facility_inventories.append(facility_inventory)

        # Bulk create all FacilityInventory entries that don't already exist
        if facility_inventories:
            FacilityInventory.objects.bulk_create(facility_inventories)

        return Response({"message": "Facility inventories created successfully."}, status=201)


class FacilityProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    search_fields = ['facility', 'product_variant']
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = ProductFilter  # keep your existing FilterSet; see example below if you don’t have one

    # Search will work on these text fields (FKs use the related field paths)
    search_fields = [
        'name', 'description', 'slug', 'tags',
        'category__name', 'brand__name',
    ]

    # Allow ordering by these fields + default ordering
    ordering_fields = [
        'name', 'category__name', 'brand__name', 'is_active', 'created_date', 'updation_date'
    ]
    ordering = ['-updation_date']
    
    def _manager_facility(self, user):
        """Retrieve the facility managed by the user (manager role)."""
        return Facility.objects.filter(managers=user).first()

    def get_queryset(self):
        user = self.request.user

        # If the user is a manager, filter the products by the facility they manage
        if getattr(user, "role", "") == "manager":
            managed_facility = self._manager_facility(user)
            if managed_facility:
                # Prefetch only the variants assigned to the manager's facility
                return Product.objects.prefetch_related(
                    Prefetch(
                        'variants',  # Prefetch related variants
                        queryset=ProductVariant.objects.filter(
                            facility_inventories__facility=managed_facility
                        ).distinct(),
                        to_attr='assigned_variants'  # Store prefetched variants in the assigned_variants attribute
                    )
                ).filter(
                    variants__facility_inventories__facility=managed_facility
                ).distinct()
            else:
                # If manager has no facility, return empty queryset
                return Product.objects.none()
        else:
            # For other users, return all products (or adjust as needed)
            return Product.objects.all()

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return ProductListSerializer
        return FacilityInventoryItemSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        
        if getattr(self.request.user, "role", "") == "manager":
            managed_facility = self._manager_facility(self.request.user)
            if managed_facility:
                ctx["facility_scope"] = [managed_facility.id]
        
        elif getattr(self.request.user, "role", "") == "facility_user":
            facility = self._facility_for_user(self.request.user)
            if facility:
                ctx["facility_scope"] = [facility.id]
        
        return ctx
    
    @action(detail=True, methods=["put"], url_path="update")
    def update_inventories(self, request, pk=None):
        """
        PUT /api/facilityproducts/{product_id}/update/
        Body: [
          { "facility_id": 1, "product_variant_id": 1, "price": 99.99, "stock": 50, "is_active": true },
          { "facility_id": 1, "product_variant_id": 2, "price": 149.99, "stock": 40, "is_active": true }
        ]
        """
        # Validate input data
        serializer = FacilityInventoryItemSerializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        items = serializer.validated_data

        # Ensure variants belong to this product
        variant_ids = {item["product_variant_id"] for item in items}
        valid_variant_ids = set(
            ProductVariant.objects.filter(product_id=pk, id__in=variant_ids)
            .values_list("id", flat=True)
        )
        invalid_variant_ids = variant_ids - valid_variant_ids
        if invalid_variant_ids:
            return Response(
                {"error": f"These variants do not belong to product {pk}: {list(invalid_variant_ids)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = []
        updated = []
        user = request.user
        # Determine facility_id from user if possible
        user_facility_id = None
        if getattr(user, "role", "") == "manager":
            managed_facility = Facility.objects.filter(managers=user).first()
            if managed_facility:
                user_facility_id = managed_facility.id

        for item in items:
            # Use facility_id from user if available, else from item
            facility_id = user_facility_id if user_facility_id else item["facility_id"]
            product_variant_id = item["product_variant_id"]

            # Check if the inventory exists for the facility and variant
            inventory = FacilityInventory.objects.filter(
                facility_id=facility_id,
                product_variant_id=product_variant_id
            ).first()

            if inventory:
                # Update the existing inventory with new values
                for field in ["base_price", "mrp", "selling_price", "stock", "cust_discount", "max_purchase_limit",
                              "outofstock_threshold", "status", "is_active", "tax_id"]:
                    if field in item:
                        setattr(inventory, field, item[field])
                inventory.save()
                updated.append(inventory.id)
            else:
                # Create a new inventory row if it does not exist
                item_with_facility = dict(item)
                item_with_facility["facility_id"] = facility_id
                new_inventory = FacilityInventory.objects.create(**item_with_facility)
                created.append(new_inventory.id)

        return Response(
            {
                "message": "Facility product updated successfully",
                "product_id": pk,
                "updated": len(updated),
                "created": len(created),
                "total": len(updated) + len(created),
            },
            status=status.HTTP_200_OK,
        )


class ClusterExportView(APIView):
    """
    GET /api/clusters/export/
    Exports clusters data in Excel format.

    Query Parameters:
    - search: Search in cluster name
    - status: Filter by active/inactive status
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum records (default: 1000, max: 10000)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        from cms.utils.filter import ClusterFilter

        # Apply filters using the same filter class as ClusterViewSet
        filter_instance = ClusterFilter(request.GET, queryset=Cluster.objects.prefetch_related('facilities', 'facilities__managers'))
        queryset = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query)
            )

        # Apply ordering same as ClusterViewSet
        queryset = queryset.order_by('name')

        # Limit results for performance
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)
        queryset = queryset[:export_limit]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clusters"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'ID', 'Name', 'Status', 'Total Facilities',
            'Total Managers', 'Created Date', 'Updated Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, cluster in enumerate(queryset, 2):
            facility_count = cluster.facilities.count()
            manager_count = sum(facility.managers.count() for facility in cluster.facilities.all())

            data = [
                cluster.id,
                cluster.name,
                'Active' if cluster.is_active else 'Inactive',
                facility_count,
                manager_count,
                cluster.creation_date.strftime('%Y-%m-%d %H:%M:%S') if cluster.creation_date else '',
                cluster.updation_date.strftime('%Y-%m-%d %H:%M:%S') if cluster.updation_date else ''
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [1, 5, 6]:  # ID, Facility count, Manager count
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="clusters_export_{timestamp}.xlsx"'

        wb.save(response)
        return response


class FacilityExportView(APIView):
    """
    GET /api/facilities/export/
    Exports facilities data in Excel format with managers and inventory info.

    Query Parameters:
    - search: Search in facility name or address
    - status: Filter by active/inactive status
    - cluster: Filter by cluster ID
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum records (default: 1000, max: 10000)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        from cms.utils.filter import FacilityFilter
        from django.db import models

        # Apply filters using the same filter class as FacilityViewSet
        # Optimize with prefetch_related for related objects and annotate for counts
        queryset = Facility.objects.prefetch_related(
            'managers',   # Prefetch managers
            'clusters',   # Prefetch clusters
        ).annotate(
            manager_count=models.Count('managers', distinct=True),
            inventory_count=models.Count('facility_inventories', distinct=True)
        )

        filter_instance = FacilityFilter(request.GET, queryset=queryset)
        queryset = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(address__icontains=search_query)
            )

        # Apply ordering same as FacilityViewSet
        queryset = queryset.order_by('name')

        # Limit results for performance
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)
        queryset = queryset[:export_limit]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Facilities"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'ID', 'Name', 'Facility Type', 'Cluster', 'Status', 'Address', 'City', 'State', 'Country',
            'Pincode', 'Latitude', 'Longitude', 'Servicable Area', 'Manager Usernames', 'Total Managers',
            'Inventory Items', 'Created Date', 'Updated Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, facility in enumerate(queryset, 2):
            # Use annotated counts instead of database queries
            manager_count = facility.manager_count
            inventory_count = facility.inventory_count

            # Get cluster names using prefetched data
            cluster_names = ', '.join([cluster.name for cluster in facility.clusters.all()]) or 'No Cluster'

            # Get manager usernames using prefetched data
            manager_usernames = ', '.join([manager.username for manager in facility.managers.all()]) or 'No Managers'

            # Handle servicable_area (could be list or text)
            servicable_area = facility.servicable_area
            if isinstance(servicable_area, list):
                servicable_area = ', '.join(servicable_area) if servicable_area else ''
            elif servicable_area is None:
                servicable_area = ''

            data = [
                facility.id,
                facility.name,
                facility.get_facility_type_display() if hasattr(facility, 'get_facility_type_display') else facility.facility_type,
                cluster_names,
                'Active' if facility.is_active else 'Inactive',
                facility.address or '',
                facility.city or '',
                facility.state or '',
                facility.country or '',
                facility.pincode or '',
                facility.latitude or '',
                facility.longitude or '',
                servicable_area,
                manager_usernames,
                manager_count,
                inventory_count,
                facility.creation_date.strftime('%Y-%m-%d %H:%M:%S') if facility.creation_date else '',
                facility.updation_date.strftime('%Y-%m-%d %H:%M:%S') if facility.updation_date else ''
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [1, 15, 16]:  # ID, Total Managers, Inventory Items
                    cell.alignment = Alignment(horizontal='center')

        # Set fixed column widths for better performance
        column_widths = [8, 20, 15, 20, 10, 30, 15, 15, 15, 10, 12, 12, 25, 25, 12, 12, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="facilities_export_{timestamp}.xlsx"'

        wb.save(response)
        return response