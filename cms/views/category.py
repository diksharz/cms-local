from rest_framework import viewsets, status
from django.db import models
from cms.models.category import Category, Brand
from cms.serializers.category import CategorySerializer, CategoryListSerializer, BrandSerializer, CategoryShelfLifeBulkUpdateSerializer
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions, AllowAny
from rest_framework.filters import SearchFilter, OrderingFilter
from cms.utils.pagination import CustomPageNumberPagination
from cms.utils.filter import BrandFilter, CategoryFilter
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from django.db import transaction
from django_filters import rest_framework as filters
from django.http import HttpResponse
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

class CategoryViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows categories to be viewed, created, updated, or deleted.
    """
    queryset = Category.objects.all()
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = CustomPageNumberPagination
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_class = CategoryFilter
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'rank', 'is_active', 'creation_date', 'updation_date']
    ordering = ['rank', 'name']

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return CategoryListSerializer
        else:
            return CategorySerializer

    def update(self, request, *args, **kwargs):
        """
        Custom update method with rank handling
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Get the requested rank from the request
        requested_rank = request.data.get('rank')
        parent_id = request.data.get('parent', instance.parent_id)
        
        # Validate serializer first
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        # Handle rank logic
        with transaction.atomic():
            if requested_rank is not None:
                # User provided a specific rank
                self._handle_rank_assignment(instance, parent_id, requested_rank)
            else:
                # No rank provided, assign automatically
                self._assign_automatic_rank(instance, parent_id)
            
            # Save the instance
            self.perform_update(serializer)
            
            # Fix any duplicate ranks after the operation
            self._fix_duplicate_ranks(parent_id)
        
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Custom create method with rank handling
        """
        # Get the requested rank from the request
        requested_rank = request.data.get('rank')
        parent_id = request.data.get('parent')
        
        # Validate serializer first
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Handle rank logic
        with transaction.atomic():
            # Create the instance first
            instance = serializer.save()
            
            if requested_rank is not None:
                # User provided a specific rank
                self._handle_rank_assignment(instance, parent_id, requested_rank)
                instance.save()
            else:
                # No rank provided, assign automatically
                self._assign_automatic_rank(instance, parent_id)
                instance.save()
            
            # Fix any duplicate ranks after the operation
            self._fix_duplicate_ranks(parent_id)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def _fix_duplicate_ranks(self, parent_id):
        """
        Fix any duplicate ranks within a parent category
        """
        siblings = Category.objects.filter(parent_id=parent_id).order_by('rank', 'id')
        
        for index, category in enumerate(siblings):
            if category.rank != index:
                category.rank = index
                category.save(update_fields=['rank'])

    def _handle_rank_assignment(self, instance, parent_id, requested_rank):
        """
        Handle rank assignment with direct swap logic
        When moving to a position, directly swap with the category at that position
        """
        # Get all siblings (categories with same parent)
        siblings = Category.objects.filter(parent_id=parent_id).exclude(id=instance.id)
        
        # If no siblings, rank is 0
        if not siblings.exists():
            instance.rank = 0
            return
        
        # Get current rank of the instance
        current_rank = instance.rank
        
        # Find the maximum rank among siblings
        max_rank = siblings.aggregate(max_rank=models.Max('rank'))['max_rank'] or 0
        
        # If requested rank is higher than max, set it to max + 1
        if requested_rank > max_rank:
            instance.rank = max_rank + 1
            return
        
        # Handle direct swap logic
        if current_rank < requested_rank:
            # Moving down: 0→1, 1→0 OR 0→3, 3→2, 2→1, 1→0
            # Shift all categories from current_rank+1 to requested_rank down by 1
            siblings.filter(
                rank__gt=current_rank,
                rank__lte=requested_rank
            ).update(rank=models.F('rank') - 1)
            
        elif current_rank > requested_rank:
            # Moving up: 1→0, 0→1 OR 3→0, 0→1, 1→2, 2→3
            # Shift all categories from requested_rank to current_rank-1 up by 1
            siblings.filter(
                rank__gte=requested_rank,
                rank__lt=current_rank
            ).update(rank=models.F('rank') + 1)
            
        else:
            # Same rank - no change needed
            return
        
        # Set the new rank for the instance
        instance.rank = requested_rank

    def _assign_automatic_rank(self, instance, parent_id):
        """
        Assign automatic rank when no rank is provided
        """
        # Get all siblings (categories with same parent)
        siblings = Category.objects.filter(parent_id=parent_id).exclude(id=instance.id)
        
        if not siblings.exists():
            # No siblings, rank is 0
            instance.rank = 0
        else:
            # Get the highest rank and add 1
            max_rank = siblings.aggregate(max_rank=models.Max('rank'))['max_rank'] or 0
            instance.rank = max_rank + 1

    @action(detail=False, methods=["get"], url_path="list")
    def tree(self, request):
        """
        Return categories as a nested tree.
        Supports:
          ?is_active=true/false  (if omitted -> no active filter)
          ?parent=<id>           (optional: limit to a parent's direct children tree)
          ?search=<term>         (includes ancestors so hierarchy is intact)
          ?page=<n>&page_size=<n>  (pagination over root nodes)
        """
        active = (request.query_params.get("is_active") or "").lower()
        parent_id = (request.query_params.get("parent") or "")
        search = (request.query_params.get("search") or "").strip()

        qs = self.get_queryset()

        # Apply active filter ONLY if explicitly provided
        if active == "true":
            qs = qs.filter(is_active=True)
        elif active == "false":
            qs = qs.filter(is_active=False)

        if parent_id:
            qs = qs.filter(parent_id=parent_id)

        base_qs = qs  # keep after active/parent filters

        # --- search with ancestors ---
        if search:
            matched_ids = list(base_qs.filter(name__icontains=search).values_list("id", flat=True))
            # Return an empty paginated result if nothing matched
            if not matched_ids:
                paginator = getattr(self, "paginator", None) or PageNumberPagination()
                if not getattr(self, "paginator", None):
                    # fallback default page size if no pagination_class on viewset
                    try:
                        paginator.page_size = int(request.query_params.get("page_size", 10))
                    except Exception:
                        paginator.page_size = 10
                page = paginator.paginate_queryset([], request, view=self)
                return paginator.get_paginated_response(page)

            parent_map = dict(base_qs.values_list("id", "parent_id"))

            final_ids = set(matched_ids)
            for mid in matched_ids:
                pid = parent_map.get(mid)
                while pid is not None and pid not in final_ids:
                    final_ids.add(pid)
                    pid = parent_map.get(pid)

            qs = base_qs.filter(id__in=final_ids)

        qs = qs.only("id", "name", "description", "parent_id", "image", "is_active", "rank") \
               .select_related("parent").order_by("rank", "name")

        # build dict of nodes
        nodes = {
            c.id: {
                "id": c.id,
                "name": c.name,
                "description": c.description,
                "parent": c.parent_id,
                "image": c.image,
                "is_active": c.is_active,
                "rank": c.rank,
                "children": [],
            }
            for c in qs
        }

        # stitch children & collect roots
        roots = []
        for c in qs:
            node = nodes[c.id]
            if c.parent_id and c.parent_id in nodes:
                nodes[c.parent_id]["children"].append(node)
            else:
                roots.append(node)

        # sort siblings recursively by rank, then by name
        def sort_children(n):
            n["children"].sort(key=lambda x: (x["rank"], x["name"].lower()))
            for ch in n["children"]:
                sort_children(ch)

        for r in roots:
            sort_children(r)

        # ---- paginate roots ----
        paginator = getattr(self, "paginator", None) or PageNumberPagination()
        if not getattr(self, "paginator", None):
            # fallback page size if no pagination_class on the viewset
            try:
                paginator.page_size = int(request.query_params.get("page_size", 10))
            except Exception:
                paginator.page_size = 10

        page = paginator.paginate_queryset(roots, request, view=self)
        return paginator.get_paginated_response(page)

    @action(detail=False, methods=["post"], url_path="bulk-shelf-life-update")
    def bulk_shelf_life_update(self, request):
        """
        Bulk update shelf life requirements for multiple categories.
        Expected payload:
        {
            "categories": [
                {"id": 1, "shelf_life_required": true},
                {"id": 2, "shelf_life_required": false}
            ]
        }
        """
        serializer = CategoryShelfLifeBulkUpdateSerializer(data=request.data)
        if serializer.is_valid():
            updated_categories = serializer.save()
            response_data = CategoryListSerializer(updated_categories, many=True).data
            return Response({
                "message": f"Successfully updated shelf life requirements for {len(updated_categories)} categories",
                "updated_categories": response_data
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class SubcategoryViewSet(viewsets.ModelViewSet):
#     """
#     API endpoint that allows subcategories to be viewed, created, updated, or deleted.
#     """
#     queryset = Subcategory.objects.all()
#     serializer_class = SubcategorySerializer
#     permission_classes = [IsAuthenticated, DjangoModelPermissions]
#     pagination_class = CustomPageNumberPagination  # Add pagination class here
#     filter_backends = (SearchFilter,)  # Add SearchFilter
#     search_fields = ['name']

# class SubsubcategoryViewSet(viewsets.ModelViewSet):
#     """
#     API endpoint that allows subsubcategories to be viewed, created, updated, or deleted.
#     """
#     queryset = Subsubcategory.objects.all()
#     serializer_class = SubsubcategorySerializer
#     permission_classes = [IsAuthenticated, DjangoModelPermissions]
#     pagination_class = CustomPageNumberPagination  # Add pagination class here
#     filter_backends = (SearchFilter,)  # Add SearchFilter
#     search_fields = ['name']

class BrandViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows brands to be viewed, created, updated, or deleted.
    """
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [AllowAny]  # Temporarily allow for testing
    pagination_class = CustomPageNumberPagination  # Add pagination class here
    filter_backends = (filters.DjangoFilterBackend, SearchFilter, OrderingFilter)  # Add SearchFilter
    filterset_class = BrandFilter
    search_fields = ['name']
    ordering_fields = ['name', 'is_active', 'creation_date', 'updation_date']
    ordering = ['name']

    def get_queryset(self):
        """
        Annotate queryset with variant count for each brand.
        """
        from django.db.models import Count
        from cms.models.product import ProductVariant
        
        return Brand.objects.annotate(
            variant_count=Count('products__variants', distinct=True)
        )

    def list(self, request, *args, **kwargs):
        from django.db.models import Q
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculate counts from filtered queryset (respects search and filters)
        active_count = queryset.filter(is_active=True).count()
        inactive_count = queryset.filter(is_active=False).count()
        # Count brands with non-null and non-empty images from filtered results
        brands_with_images_count = queryset.exclude(Q(image='') | Q(image__isnull=True)).count()
        
        # Paginate
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data['total_active_count'] = active_count
            response.data['total_inactive_count'] = inactive_count
            response.data['total_brands_with_images_count'] = brands_with_images_count
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'count': queryset.count(),
            'total_active_count': active_count,
            'total_inactive_count': inactive_count,
            'total_brands_with_images_count': brands_with_images_count
        })


class BrandExportView(APIView):
    """
    GET /api/brands/export/
    Exports brands data in Excel format.

    Query Parameters:
    - search: Search in brand name
    - status: Filter by active/inactive status
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum records (default: 1000, max: 10000)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        # Apply filters using the same filter class as BrandViewSet
        filter_instance = BrandFilter(request.GET, queryset=Brand.objects.all())
        queryset = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | Q(description__icontains=search_query)
            )

        # Apply ordering same as BrandViewSet
        queryset = queryset.order_by('name')

        # Limit results for performance
        export_limit = min(int(request.GET.get('limit', 1000)), 10000)
        queryset = queryset[:export_limit]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Brands"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'ID', 'Name', 'Description', 'Status', 'Has Image',
            'Created Date', 'Updated Date', 'Total Products'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, brand in enumerate(queryset, 2):
            product_count = brand.products.count()

            data = [
                brand.id,
                brand.name,
                brand.description or '',
                'Active' if brand.is_active else 'Inactive',
                'Yes' if brand.image else 'No',
                brand.creation_date.strftime('%Y-%m-%d %H:%M:%S') if brand.creation_date else '',
                brand.updation_date.strftime('%Y-%m-%d %H:%M:%S') if brand.updation_date else '',
                product_count
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [1, 8]:  # ID and Total Products columns
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="brands_export_{timestamp}.xlsx"'

        wb.save(response)
        return response


class CategoryExportView(APIView):
    """
    GET /api/categories/export/
    Exports categories data in Excel format with hierarchy.

    Query Parameters:
    - search: Search in category name
    - status: Filter by active/inactive status
    - format: Export format (excel/csv) - default: excel
    - limit: Maximum records (default: 1000, max: 10000)
    """
    permission_classes = [AllowAny]

    def get(self, request):
        # Get all categories with parent-child relationships
        all_categories = Category.objects.select_related('parent').prefetch_related('children', 'products').annotate(
            product_count=models.Count('products', distinct=True),
            subcategory_count=models.Count('children', distinct=True)
        )

        # Apply filters
        filter_instance = CategoryFilter(request.GET, queryset=all_categories)
        filtered_categories = filter_instance.qs

        # Apply search if provided
        search_query = request.GET.get('search', '')
        if search_query:
            filtered_categories = filtered_categories.filter(
                Q(name__icontains=search_query) | Q(description__icontains=search_query)
            )

        # Build hierarchical tree structure
        def build_tree_data(categories):
            tree_data = []
            processed = set()

            def add_category_with_children(category, level=0):
                if category.id in processed:
                    return
                processed.add(category.id)

                # Add indentation based on level
                indent = "  " * level
                tree_data.append({
                    'category': category,
                    'level': level,
                    'indent': indent
                })

                # Add children recursively
                children = sorted([child for child in categories if child.parent_id == category.id], key=lambda x: (x.rank, x.name))
                for child in children:
                    add_category_with_children(child, level + 1)

            # Start with root categories (no parent)
            root_categories = sorted([cat for cat in categories if cat.parent_id is None], key=lambda x: (x.rank, x.name))
            for root in root_categories:
                add_category_with_children(root)

            return tree_data

        tree_data = build_tree_data(list(filtered_categories))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories Hierarchy"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        root_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        level1_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        level2_fill = PatternFill(start_color="F8FCFF", end_color="F8FCFF", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Level', 'Category Name', 'ID', 'Description', 'Status', 'Rank',
            'Shelf Life Required', 'Products', 'Subcategories', 'Created Date', 'Updated Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows with hierarchy
        for row_idx, item in enumerate(tree_data, 2):
            category = item['category']
            level = item['level']
            indent = item['indent']

            # Choose background color based on level
            if level == 0:
                bg_fill = root_fill
                font_style = Font(bold=True)
            elif level == 1:
                bg_fill = level1_fill
                font_style = Font(bold=False)
            else:
                bg_fill = level2_fill
                font_style = Font(bold=False)

            data = [
                f"Level {level}",
                f"{indent}{category.name}",
                category.id,
                category.description or '',
                'Active' if category.is_active else 'Inactive',
                category.rank,
                'Yes' if category.shelf_life_required else 'No',
                category.product_count,
                category.subcategory_count,
                category.creation_date.strftime('%Y-%m-%d %H:%M:%S') if category.creation_date else '',
                category.updation_date.strftime('%Y-%m-%d %H:%M:%S') if category.updation_date else ''
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.fill = bg_fill
                if col == 2:  # Category name column
                    cell.font = font_style
                if col in [3, 6, 8, 9]:  # ID, Rank, Products, Subcategories
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="categories_export_{timestamp}.xlsx"'

        wb.save(response)
        return response
